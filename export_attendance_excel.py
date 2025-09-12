#!/usr/bin/env python3
"""
Enhanced Excel Export Tool for Attendance Data with Late Clock-In Highlighting
Highlights late clock-ins in red for easy identification
"""

import sys
from datetime import datetime, timedelta
from core.mongodb_manager import MongoDBManager

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
    print("   Falling back to CSV export...")

def export_attendance_to_excel(days=30, filename=None):
    """
    Export attendance records to Excel with late clock-in highlighting
    """
    if not EXCEL_AVAILABLE:
        print("Excel export not available. Please install openpyxl.")
        return export_attendance_to_csv_fallback(days, filename)
    
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_with_highlighting_{timestamp}.xlsx"
    
    db = MongoDBManager()
    
    try:
        print(f"Exporting attendance data for last {days} days to Excel...")
        
        # Get attendance data for specified period
        start_date = datetime.now() - timedelta(days=days)
        end_date = datetime.now()
        
        # MongoDB aggregation pipeline
        pipeline = [
            {
                "$match": {
                    "timestamp": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$lookup": {
                    "from": "employees",
                    "localField": "employee_id",
                    "foreignField": "employee_id",
                    "as": "employee"
                }
            },
            {
                "$unwind": "$employee"
            },
            {
                "$sort": {"timestamp": -1}
            },
            {
                "$project": {
                    "employee_id": 1,
                    "name": "$employee.name",
                    "department": "$employee.department",
                    "date": {"$dateToString": {"date": "$timestamp", "format": "%Y-%m-%d"}},
                    "time": {"$dateToString": {"date": "$timestamp", "format": "%H:%M:%S"}},
                    "timestamp": {"$dateToString": {"date": "$timestamp", "format": "%Y-%m-%d %H:%M:%S"}},
                    "method": 1,
                    "status": 1,
                    "attendance_type": 1,
                    "late": {"$ifNull": ["$late", False]},
                    "location_name": {"$ifNull": ["$location_name", ""]},
                    "address": {"$ifNull": ["$address", ""]}
                }
            }
        ]
        
        records = list(db.attendance.aggregate(pipeline))
        
        if not records:
            print("No records found for the specified period")
            return
        
        # Create Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        late_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red background
        late_font = Font(color="FFFFFF", bold=True)
        
        normal_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        # Headers
        headers = [
            'Employee ID', 'Name', 'Department', 'Date', 'Time', 'Full Timestamp',
            'Method', 'Status', 'Type', 'Late', 'Location Name', 'Address'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        late_rows = []
        for row_num, record in enumerate(records, 2):
            is_late_clock_in = (
                record.get('late', False) and 
                record.get('attendance_type', '').upper() == 'CLOCK' and 
                record.get('status', '') == 'in'
            )
            
            if is_late_clock_in:
                late_rows.append(row_num)
            
            row_data = [
                record.get('employee_id', ''),
                record.get('name', ''),
                record.get('department', ''),
                record.get('date', ''),
                record.get('time', ''),
                record.get('timestamp', ''),
                record.get('method', ''),
                record.get('status', ''),
                record.get('attendance_type', ''),
                'YES' if record.get('late', False) else 'NO',
                record.get('location_name', ''),
                record.get('address', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Apply late highlighting
                if is_late_clock_in:
                    cell.fill = late_fill
                    cell.font = late_font
                else:
                    # Center align certain columns
                    if col_num in [1, 4, 5, 7, 8, 9, 10]:  # ID, Date, Time, Status, Type, Late
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = normal_alignment
        
        # Auto-adjust column widths
        column_widths = {
            1: 12,  # Employee ID
            2: 20,  # Name
            3: 15,  # Department
            4: 12,  # Date
            5: 10,  # Time
            6: 20,  # Full Timestamp
            7: 15,  # Method
            8: 8,   # Status
            9: 8,   # Type
            10: 6,  # Late
            11: 25, # Location Name
            12: 35  # Address
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add summary information
        summary_row = len(records) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value=f"Total Records: {len(records)}")
        ws.cell(row=summary_row + 2, column=1, value=f"Late Clock-ins: {len(late_rows)}")
        ws.cell(row=summary_row + 3, column=1, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if late_rows:
            ws.cell(row=summary_row + 4, column=1, value="Late clock-in rows highlighted in RED")
            ws.cell(row=summary_row + 4, column=1).font = Font(color="FF6B6B", bold=True)
        
        # Save the workbook
        wb.save(filename)
        
        print(f"✅ Exported {len(records)} records to {filename}")
        if late_rows:
            print(f"🔴 {len(late_rows)} late clock-in records highlighted in red")
        print(f"📊 Summary: {len(records)} total records, {len(late_rows)} late arrivals")
        
        return filename
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        return None

def export_attendance_to_csv_fallback(days=30, filename=None):
    """
    Fallback CSV export when openpyxl is not available
    """
    import csv
    
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_fallback_{timestamp}.csv"
    
    db = MongoDBManager()
    
    try:
        print(f"Exporting attendance data for last {days} days to CSV (fallback)...")
        
        # Get attendance data for specified period
        start_date = datetime.now() - timedelta(days=days)
        end_date = datetime.now()
        
        # MongoDB aggregation pipeline
        pipeline = [
            {
                "$match": {
                    "timestamp": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$lookup": {
                    "from": "employees",
                    "localField": "employee_id",
                    "foreignField": "employee_id",
                    "as": "employee"
                }
            },
            {
                "$unwind": "$employee"
            },
            {
                "$sort": {"timestamp": -1}
            }
        ]
        
        records = list(db.attendance.aggregate(pipeline))
        
        if not records:
            print("No records found for the specified period")
            return
        
        # CSV Headers
        headers = [
            'Employee ID', 'Name', 'Department', 'Date', 'Time', 'Timestamp',
            'Method', 'Status', 'Type', 'Late', 'Location Name', 'Address'
        ]
        
        # Write to CSV
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            late_count = 0
            for record in records:
                is_late = record.get('late', False)
                if is_late:
                    late_count += 1
                    
                row_data = [
                    record.get('employee_id', ''),
                    record.get('name', ''),
                    record.get('department', ''),
                    record.get('date', ''),
                    record.get('time', ''),
                    record.get('timestamp', ''),
                    record.get('method', ''),
                    record.get('status', ''),
                    record.get('attendance_type', ''),
                    'YES' if is_late else 'NO',
                    record.get('location_name', ''),
                    record.get('address', '')
                ]
                writer.writerow(row_data)
        
        print(f"✅ Exported {len(records)} records to {filename}")
        print(f"⚠️  {late_count} late clock-in records (cannot highlight in CSV)")
        
        return filename
        
    except Exception as e:
        print(f"❌ Error exporting to CSV: {e}")
        return None

def main():
    """Main function to run the export"""
    print("=== Attendance Export Tool ===")
    print("Export options:")
    print("1. Last 7 days")
    print("2. Last 30 days") 
    print("3. Last 90 days")
    print("4. Custom days")
    
    try:
        choice = input("\nEnter your choice (1-4): ").strip()
        
        if choice == "1":
            days = 7
        elif choice == "2":
            days = 30
        elif choice == "3":
            days = 90
        elif choice == "4":
            try:
                days = int(input("Enter number of days: "))
                if days <= 0:
                    print("Invalid number of days")
                    return
            except ValueError:
                print("Invalid input")
                return
        else:
            print("Invalid choice")
            return
        
        filename = export_attendance_to_excel(days)
        if filename:
            print(f"\n🎉 Export completed successfully!")
            print(f"📁 File saved as: {filename}")
            if EXCEL_AVAILABLE:
                print("🔴 Late clock-ins are highlighted in red in the Excel file")
        else:
            print("❌ Export failed")
            
    except KeyboardInterrupt:
        print("\n\n⏹️  Export cancelled by user")
    except Exception as e:
        print(f"\n❌ Error during export: {e}")

if __name__ == "__main__":
    main()
