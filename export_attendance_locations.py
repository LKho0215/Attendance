#!/usr/bin/env python3
"""
Excel Export Tool for Unified Attendance + Location Data
Demonstrates how easy it is to export all data from a single table
"""

import os
from datetime import datetime, timedelta
from core.mongodb_manager import MongoDBManager
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def export_attendance_to_excel(days=30, filename=None):
    """
    Export attendance records with location data to Excel
    Highlights late clock-ins in red
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exports/attendance_with_locations_{timestamp}.xlsx"
    
    # Ensure exports directory exists
    os.makedirs("exports", exist_ok=True)
    
    db = MongoDBManager()
    
    try:
        print(f"Exporting attendance data for last {days} days...")
        
        # Get attendance data for specified period
        start_date = datetime.now() - timedelta(days=days)
        end_date = datetime.now() + timedelta(days=1)  # Include future records from today
        
        # Simple aggregation - no complex joins needed!
        pipeline = [
            {
                "$match": {
                    "timestamp": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$lookup": {
                    "from": "employees",
                    "localField": "employee_id",
                    "foreignField": "employee_id",
                    "as": "employee"
                }
            },
            {
                "$addFields": {
                    "employee": {
                        "$ifNull": [
                            {"$arrayElemAt": ["$employee", 0]},
                            {
                                "name": "Unknown Employee",
                                "department": "Unknown",
                                "employee_id": "$employee_id"
                            }
                        ]
                    }
                }
            },
            {
                "$sort": {"timestamp": -1}
            },
            {
                "$project": {
                    "employee_id": 1,
                    "name": "$employee.name",
                    "department": "$employee.department",
                    "date": {"$dateToString": {"date": "$timestamp", "format": "%Y-%m-%d"}},
                    "time": {"$dateToString": {"date": "$timestamp", "format": "%H:%M:%S"}},
                    "timestamp": {"$dateToString": {"date": "$timestamp", "format": "%Y-%m-%d %H:%M:%S"}},
                    "method": 1,
                    "status": 1,
                    "attendance_type": 1,
                    "location_name": {"$ifNull": ["$location_name", ""]},
                    "address": {"$ifNull": ["$address", ""]},
                    "late": {"$ifNull": ["$late", False]}
                }
            }
        ]
        
        records = list(db.attendance.aggregate(pipeline))
        
        if not records:
            print("No records found for the specified period")
            return
        
        # Create Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Define styles
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Late clock-in style (red background)
        late_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        late_font = Font(color="CC0000", bold=True)
        
        # Regular style
        regular_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Excel Headers
        headers = [
            'Employee ID', 'Name', 'Department', 'Date', 'Time', 'Timestamp',
            'Method', 'Status', 'Type', 'Location Name', 'Address', 'Late Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data rows
        late_count = 0
        for row, record in enumerate(records, 2):  # Start from row 2 (after header)
            is_late = record.get('late', False)
            
            row_data = [
                record.get('employee_id', ''),
                record.get('name', ''),
                record.get('department', ''),
                record.get('date', ''),
                record.get('time', ''),
                record.get('timestamp', ''),
                record.get('method', ''),
                record.get('status', ''),
                record.get('attendance_type', ''),
                record.get('location_name', ''),
                record.get('address', ''),
                'LATE' if is_late else 'ON-TIME'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = regular_alignment
                cell.border = border
                
                # Highlight late clock-ins
                if is_late and record.get('attendance_type') == 'clock' and record.get('status') == 'in':
                    cell.fill = late_fill
                    if col == 12:  # Late Status column
                        cell.font = late_font
                    late_count += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        
        print(f"✅ Exported {len(records)} records to {filename}")
        
        # Count actual late clock-ins
        actual_late_count = sum(1 for r in records 
                               if r.get('late', False) and 
                                  r.get('attendance_type') == 'clock' and 
                                  r.get('status') == 'in')
        
        if actual_late_count > 0:
            print(f"🔴 {actual_late_count} late clock-in records highlighted in red")
        
        # Show sample of exported data
        print(f"\nSample of exported data:")
        print("-" * 80)
        for i, record in enumerate(records[:5]):
            location_info = record.get('location_name', 'No location')
            late_status = " (LATE)" if record.get('late', False) else ""
            print(f"{record['name']}{late_status} | {record['timestamp']} | {record['status'].upper()} | {location_info}")
        
        if len(records) > 5:
            print(f"... and {len(records) - 5} more records")
            
        return filename
        
    finally:
        db.close_connection()

def export_check_outs_only():
    """Export only CHECK OUT records with locations"""
    db = MongoDBManager()
    
    try:
        # Filter for CHECK OUT records only
        pipeline = [
            {
                "$match": {
                    "attendance_type": "check",
                    "status": "out",
                    "location_name": {"$exists": True, "$ne": ""}
                }
            },
            {
                "$lookup": {
                    "from": "employees",
                    "localField": "employee_id", 
                    "foreignField": "employee_id",
                    "as": "employee"
                }
            },
            {
                "$unwind": "$employee"
            },
            {
                "$sort": {"timestamp": -1}
            }
        ]
        
        records = list(db.attendance.aggregate(pipeline))
        
        print(f"\nCHECK OUT Records with Locations ({len(records)} found):")
        print("=" * 60)
        
        for record in records:
            timestamp = record['timestamp'].strftime("%Y-%m-%d %H:%M:%S")
            print(f"📍 {record['employee']['name']} | {timestamp}")
            print(f"   Location: {record.get('location_name', 'N/A')}")
            print(f"   Address:  {record.get('address', 'N/A')}")
            print()
    
    finally:
        db.close_connection()

def main():
    print("UNIFIED ATTENDANCE + LOCATION EXPORT TOOL (Excel Format)")
    print("=" * 60)
    
    print("1. Export all attendance data (Excel with late highlighting)")
    print("2. Show CHECK OUT locations only")
    print("3. Export last 7 days (Excel)")
    print("4. Export last 30 days (Excel)")
    
    choice = input("\nEnter your choice (1-4): ").strip()
    
    if choice == "1":
        filename = export_attendance_to_excel(days=365)  # Full year
        print(f"\n💾 Data exported to: {filename}")
        
    elif choice == "2":
        export_check_outs_only()
        
    elif choice == "3":
        filename = export_attendance_to_excel(days=7)
        print(f"\n💾 Data exported to: {filename}")
        
    elif choice == "4":
        filename = export_attendance_to_excel(days=30)
        print(f"\n💾 Data exported to: {filename}")
        
    else:
        print("Invalid choice!")

if __name__ == "__main__":
    main()
