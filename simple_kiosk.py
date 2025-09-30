#!/usr/bin/env python3
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
import cv2
from PIL import Image, ImageTk
import threading
import time
import csv
import os
import ctypes
import re
import numpy as np
import traceback
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from core.mongodb_manager import MongoDBManager
from core.deepface_recognition import DeepFaceRecognitionSystem, CameraManager
from core.barcode_scanner import BarcodeScanner
from core.attendance import AttendanceManager
from core.mongo_location_manager import MongoLocationManager
from core.attendance_ultra_light import AttendanceUltraLightDetector

# Set appearance mode and color theme2
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Configuration constants
CONFIDENCE_THRESHOLD = 0.6  # Minimum confidence required for face recognition

class SimpleKioskApp:
    def __init__(self):
        # Make application DPI-aware (must be done before creating tkinter window)
        try:
            import ctypes
            # Tell Windows this app is DPI-aware
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
            print("[INIT DEBUG] Successfully set DPI awareness")
        except:
            print("[INIT DEBUG] Could not set DPI awareness (not critical)")
        
        self.root = ctk.CTk()
        self.root.title("Attendance Kiosk")
        
        # Kiosk mode settings
        self.setup_kiosk_mode()
        
        # Initialize core components
        print("[INIT DEBUG] Initializing core components...")
        try:
            self.db = MongoDBManager()
            print("[INIT DEBUG] MongoDB manager initialized")
            
            self.attendance_manager = AttendanceManager(self.db)
            print("[INIT DEBUG] Attendance manager initialized")
            
            self.location_manager = MongoLocationManager(self.db)
            print("[INIT DEBUG] MongoDB location manager initialized")
            
        except Exception as e:
            print(f"[INIT ERROR] Failed to initialize MongoDB: {e}")
            messagebox.showerror("Database Error", 
                               f"Failed to connect to MongoDB:\n{e}\n\nPlease check your mongo_config.py file.")
            return
        
        # Face detection method configuration
        self.use_ultra_light_detection = True  # Set to True for ultra-fast detection, False for DeepFace
        
        if self.use_ultra_light_detection:
            # Initialize Ultra Light Face Detection for maximum performance
            print("[INIT DEBUG] Initializing Ultra Light Face Detection...")
            self.ultra_light_detector = AttendanceUltraLightDetector(confidence_threshold=0.7)
            print("[INIT DEBUG] Ultra Light Face Detection initialized")
            
            # Still keep face recognition for enrollment/verification if needed
            self.face_recognition = DeepFaceRecognitionSystem(detector_backend='mtcnn')
            print("[INIT DEBUG] DeepFace recognition system initialized (backup)")
        else:
            # Use traditional DeepFace system
            self.face_recognition = DeepFaceRecognitionSystem(detector_backend='mtcnn')
            print("[INIT DEBUG] DeepFace recognition system initialized")
            
            # Start background face processing
            self.face_recognition.start_background_processing()
            print("[INIT DEBUG] Background face processing started")
            
            self.ultra_light_detector = None
        
        self.camera_manager = CameraManager()
        print("[INIT DEBUG] Camera manager initialized")
        print("[INIT DEBUG] Camera manager initialized")
        
        self.barcode_scanner = BarcodeScanner()
        print("[INIT DEBUG] Barcode scanner initialized")
        
        # GUI state variables
        self.camera_active = False
        self.current_mode = "auto"  # auto, manual
        self.attendance_mode = "CLOCK"  # CLOCK or CHECK
        self.auto_timeout = 3  # seconds to show result
        self.last_history_update = 0  # for periodic updates
        
        # Camera management for registration
        self.main_camera_paused = False  # Flag to pause main camera during registration
        self.reg_camera_active = False  # Registration camera preview active
        self.reg_capture_active = False  # Registration capture process active
        
        # Create GUI
        self.create_interface()
        
        # Set initial mode appearance
        self.set_attendance_mode("CLOCK")
        
        # Load known faces
        self.load_known_faces()
        
        # Do NOT start camera automatically - start with camera OFF
        # Camera will be activated manually when employee presses "+"
        print("[INIT DEBUG] Camera initialized but NOT started - manual activation required")
        
        # Setup keyboard shortcuts
        self.setup_keyboard_shortcuts()
        
        # Start update loop
        self.update_loop()
    
    def setup_kiosk_mode(self):
        """Configure for kiosk operation"""
        # Get actual screen size (handling DPI scaling)
        try:
            import ctypes
            # Get actual screen dimensions bypassing DPI scaling
            user32 = ctypes.windll.user32
            screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
            screen_width, screen_height = screensize
            print(f"[KIOSK DEBUG] True screen dimensions (via Windows API): {screen_width}x{screen_height}")
        except:
            # Fallback to tkinter method
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            print(f"[KIOSK DEBUG] Fallback screen dimensions (via tkinter): {screen_width}x{screen_height}")
        
        # Also show what tkinter thinks the screen size is
        tk_width = self.root.winfo_screenwidth()
        tk_height = self.root.winfo_screenheight()
        print(f"[KIOSK DEBUG] Tkinter detected dimensions: {tk_width}x{tk_height}")
        
        try:
            # Method 1: Full override with true resolution
            self.root.geometry(f"{screen_width}x{screen_height}+0+0")
            self.root.attributes('-fullscreen', True)
            self.root.attributes('-topmost', True)
            self.root.overrideredirect(True)  # Remove all window decorations
            print(f"[KIOSK DEBUG] Using override redirect method with true resolution")
        except Exception as e:
            print(f"[KIOSK DEBUG] Override method failed: {e}")
            try:
                # Method 2: Standard fullscreen with true resolution
                self.root.geometry(f"{screen_width}x{screen_height}+0+0")
                self.root.attributes('-fullscreen', True)
                self.root.attributes('-topmost', True)
                self.root.state('zoomed')
                print(f"[KIOSK DEBUG] Using standard fullscreen method with true resolution")
            except Exception as e2:
                print(f"[KIOSK DEBUG] Standard fullscreen failed: {e2}")
                # Method 3: Fallback maximized
                self.root.state('zoomed')
                self.root.attributes('-topmost', True)
                print(f"[KIOSK DEBUG] Using fallback maximized method")
        
        # Force focus and update
        self.root.focus_force()
        self.root.update()
        
        # Additional focus methods for better reliability
        self.root.lift()  # Bring window to front
        self.root.grab_set()  # Capture all events
        
        # Check final size after a moment
        self.root.after(100, self.check_final_size)
        
        # Schedule aggressive focus setting after UI is ready
        self.root.after(500, self.ensure_focus)
    
    def check_final_size(self):
        """Check and report the final window size"""
        final_width = self.root.winfo_width()
        final_height = self.root.winfo_height()
        print(f"[KIOSK DEBUG] Final window size: {final_width}x{final_height}")
        
        # Check if we're getting the full resolution
        if final_width >= 1920 and final_height >= 1080:
            print(f"[KIOSK DEBUG] ✓ Successfully achieved full 1920x1080 resolution")
        else:
            print(f"[KIOSK DEBUG] ⚠ Window size is smaller than expected 1920x1080")
            print(f"[KIOSK DEBUG] This might be due to DPI scaling or window decorations")
    
    def ensure_focus(self):
        """Ensure the application has focus and can receive keyboard input"""
        print("[FOCUS DEBUG] Ensuring application has focus...")
        try:
            # Step 1: Use Windows API to force the window to foreground
            self.force_window_to_foreground()
            
            # Step 2: Standard tkinter focus methods
            self.root.lift()                    # Bring to front
            self.root.attributes('-topmost', True)  # Stay on top
            self.root.focus_force()             # Force focus
            self.root.focus_set()               # Set focus
            
            # Step 3: Update to apply changes
            self.root.update()
            self.root.update_idletasks()
            
            # Step 4: Test keyboard focus
            self.root.bind('<KeyPress>', self.test_key_handler, add='+')
            
            print("[FOCUS DEBUG] ✓ Focus management completed")
            
            # Show visual indication that focus test is active
            self.status_label.configure(
                text="● PRESS ANY KEY to confirm keyboard focus is working",
                text_color="yellow"
            )
            
            # Schedule a focus check in a few seconds
            self.root.after(5000, self.check_focus_status)
            
        except Exception as e:
            print(f"[FOCUS DEBUG] ✗ Focus management failed: {e}")
    
    def force_window_to_foreground(self):
        """Use Windows API to force window to foreground"""
        try:
            import ctypes
            from ctypes import wintypes
            
            # Get the window handle
            hwnd = self.root.winfo_id()
            
            # Windows API functions
            user32 = ctypes.windll.user32
            kernel32 = ctypes.windll.kernel32
            
            # Get current foreground window and thread
            current_foreground = user32.GetForegroundWindow()
            current_thread = kernel32.GetCurrentThreadId()
            foreground_thread = user32.GetWindowThreadProcessId(current_foreground, None)
            
            # Attach to the foreground thread to gain focus rights
            if foreground_thread != current_thread:
                user32.AttachThreadInput(foreground_thread, current_thread, True)
                user32.SetForegroundWindow(hwnd)
                user32.AttachThreadInput(foreground_thread, current_thread, False)
            else:
                user32.SetForegroundWindow(hwnd)
            
            # Additional focus calls
            user32.SetActiveWindow(hwnd)
            user32.SetFocus(hwnd)
            user32.ShowWindow(hwnd, 1)  # SW_NORMAL
            user32.BringWindowToTop(hwnd)
            
            print("[FOCUS DEBUG] ✓ Windows API foreground focus applied")
            return True
            
        except Exception as e:
            print(f"[FOCUS DEBUG] Windows API focus failed: {e}")
            return False
    
    def test_key_handler(self, event):
        """Test handler to verify keyboard focus is working"""
        print(f"[FOCUS DEBUG] ✓ Keyboard focus confirmed - received key: {event.keysym}")
        # Update status to show focus is working
        self.status_label.configure(
            text="● FOCUS CONFIRMED - Keyboard input is working!",
            text_color="cyan"
        )
        # Remove the test handler after first key press
        self.root.unbind('<KeyPress>')
        # Return to normal status after 3 seconds
        self.root.after(3000, lambda: self.status_label.configure(
            text="● READY - Please position yourself in front of the camera",
            text_color="green"
        ))
    
    def check_focus_status(self):
        """Check and report current focus status"""
        try:
            focused_widget = self.root.focus_get()
            if focused_widget:
                print(f"[FOCUS DEBUG] ✓ Focus is on: {focused_widget}")
                # If we still have the test message, revert to normal
                if "PRESS ANY KEY" in self.status_label.cget("text"):
                    self.status_label.configure(
                        text="● READY - Please position yourself in front of the camera",
                        text_color="green"
                    )
                    print("[FOCUS DEBUG] No key pressed during test, but focus appears to be working")
            else:
                print("[FOCUS DEBUG] ⚠ No widget has focus - reapplying focus")
                self.root.focus_force()
                self.root.focus_set()
                self.status_label.configure(
                    text="● FOCUS ISSUE - Click on window if keyboard doesn't work",
                    text_color="orange"
                )
        except Exception as e:
            print(f"[FOCUS DEBUG] Focus check failed: {e}")
    
    def maintain_focus(self):
        """Periodic focus maintenance to ensure keyboard input works"""
        try:
            # Check if we still have focus
            focused_widget = self.root.focus_get()
            if not focused_widget or focused_widget != self.root:
                # We've lost focus, regain it
                print("[FOCUS DEBUG] Focus lost, regaining...")
                self.root.lift()
                self.root.focus_force()
                self.root.focus_set()
                
                # Try Windows API if available
                try:
                    import ctypes
                    hwnd = self.root.winfo_id()
                    ctypes.windll.user32.SetForegroundWindow(hwnd)
                except:
                    pass  # Not critical if this fails
                    
        except Exception as e:
            # Don't spam the console, just try to regain focus
            self.root.focus_force()
    
    def create_interface(self):
        """Create the main interface"""
        # Main container
        self.main_frame = ctk.CTkFrame(self.root, corner_radius=0)
        self.main_frame.pack(fill="both", expand=True)
        
        # Header
        self.create_header()
        
        # Content area with camera and controls
        self.content_frame = ctk.CTkFrame(self.main_frame)
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Left side - Camera feed (smaller, fixed width)
        self.camera_frame = ctk.CTkFrame(self.content_frame, width=480, height=400)
        self.camera_frame.pack(side="left", fill="y", padx=(10, 20), pady=10)
        self.camera_frame.pack_propagate(False)  # Maintain fixed size
        
        self.camera_label = ctk.CTkLabel(
            self.camera_frame, 
            text="📷 Camera Off\n\nPress + (Plus) or Numpad +\nto activate camera for recognition\n\nManual Recognition Mode",
            font=ctk.CTkFont(size=20)
        )
        self.camera_label.pack(expand=True, padx=15, pady=15)
        
        # Add employee registration button below camera
        self.register_btn = ctk.CTkButton(
            self.camera_frame,
            text="👤 REGISTER NEW EMPLOYEE (R)",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=420,
            height=40,
            fg_color=("blue", "navy"),
            hover_color=("darkblue", "lightblue"),
            command=self.show_employee_registration_dialog
        )
        self.register_btn.pack(side="bottom", padx=15, pady=15)
        
        # Right side - Attendance history (larger, expandable)
        self.control_frame = ctk.CTkFrame(self.content_frame)
        self.control_frame.pack(side="right", fill="both", expand=True, padx=(20, 10), pady=10)
        
        self.create_controls()
        
        # Bottom - Status and messages
        self.create_status_area()
    
    def create_header(self):
        """Create header with title and time"""
        header_frame = ctk.CTkFrame(self.main_frame, height=120)
        header_frame.pack(fill="x", padx=30, pady=20)
        header_frame.pack_propagate(False)
        
        # Title
        title_label = ctk.CTkLabel(
            header_frame,
            text="ATTENDANCE SYSTEM",
            font=ctk.CTkFont(size=42, weight="bold")
        )
        title_label.pack(side="left", padx=30, pady=30)
        
        # Time display
        self.time_label = ctk.CTkLabel(
            header_frame,
            text="",
            font=ctk.CTkFont(size=24)
        )
        self.time_label.pack(side="right", padx=30, pady=30)
    
    def create_controls(self):
        """Create control panels with separate interfaces for CLOCK and CHECK modes"""
        # Mode selection frame at top
        mode_selection_frame = ctk.CTkFrame(self.control_frame, height=80)
        mode_selection_frame.pack(fill="x", padx=10, pady=10)
        mode_selection_frame.pack_propagate(False)
        
        mode_title = ctk.CTkLabel(
            mode_selection_frame,
            text="SELECT MODE:",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        mode_title.pack(side="left", padx=20, pady=20)
        
        # Mode buttons
        self.clock_btn = ctk.CTkButton(
            mode_selection_frame,
            text="🕐 CLOCK MODE (1)",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=180,
            height=50,
            command=lambda: self.set_attendance_mode("CLOCK")
        )
        self.clock_btn.pack(side="left", padx=10, pady=15)
        
        self.check_btn = ctk.CTkButton(
            mode_selection_frame,
            text="✅ CHECK MODE (2)",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=180,
            height=50,
            command=lambda: self.set_attendance_mode("CHECK")
        )
        self.check_btn.pack(side="left", padx=10, pady=15)
        
        # Create interface panels (initially hidden)
        self.create_clock_interface()
        self.create_check_interface()
        
        # Load initial attendance data
        self.update_attendance_history()
    
    def create_clock_interface(self):
        """Create the CLOCK mode interface for shift management"""
        self.clock_interface = ctk.CTkFrame(self.control_frame)
        # Don't pack it yet - it will be shown when CLOCK mode is selected
        
        # CLOCK mode title
        clock_title_frame = ctk.CTkFrame(self.clock_interface, fg_color="transparent")
        clock_title_frame.pack(fill="x", padx=10, pady=10)
        
        clock_title = ctk.CTkLabel(
            clock_title_frame,
            text="🕐 SHIFT MANAGEMENT",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=("green", "lightgreen")
        )
        clock_title.pack()
        
        # CLOCK mode attendance history
        self.clock_history_frame = ctk.CTkScrollableFrame(
            self.clock_interface,
            width=600,
            height=520,
            fg_color=("gray95", "gray10"),
            border_width=2,
            border_color=("green", "darkgreen")
        )
        self.clock_history_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # CLOCK mode controls
        clock_controls = ctk.CTkFrame(self.clock_interface, height=80)
        clock_controls.pack(fill="x", padx=10, pady=10)
        clock_controls.pack_propagate(False)
        
        self.clock_manual_btn = ctk.CTkButton(
            clock_controls,
            text="⌨️ MANUAL CLOCK ENTRY (Enter)",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=250,
            height=50,
            command=self.show_manual_entry
        )
        self.clock_manual_btn.pack(side="left", padx=10, pady=15)
        
        self.clock_status_label = ctk.CTkLabel(
            clock_controls,
            text="Scan face or QR code to CLOCK IN/OUT",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="green"
        )
        self.clock_status_label.pack(side="right", padx=20, pady=15)
    
    def create_check_interface(self):
        """Create the CHECK mode interface for office entry/exit"""
        self.check_interface = ctk.CTkFrame(self.control_frame)
        # Don't pack it yet - it will be shown when CHECK mode is selected
        
        # CHECK mode title
        check_title_frame = ctk.CTkFrame(self.check_interface, fg_color="transparent")
        check_title_frame.pack(fill="x", padx=10, pady=10)
        
        check_title = ctk.CTkLabel(
            check_title_frame,
            text="✅ OFFICE ENTRY/EXIT",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=("blue", "lightblue")
        )
        check_title.pack()
        
        # CHECK mode attendance history
        self.check_history_frame = ctk.CTkScrollableFrame(
            self.check_interface,
            width=600,
            height=520,
            fg_color=("gray95", "gray10"),
            border_width=2,
            border_color=("blue", "darkblue")
        )
        self.check_history_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # CHECK mode controls
        check_controls = ctk.CTkFrame(self.check_interface, height=80)
        check_controls.pack(fill="x", padx=10, pady=10)
        check_controls.pack_propagate(False)
        
        self.check_manual_btn = ctk.CTkButton(
            check_controls,
            text="⌨️ MANUAL CHECK ENTRY (Enter)",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=250,
            height=50,
            command=self.show_manual_entry
        )
        self.check_manual_btn.pack(side="left", padx=10, pady=15)
        
        self.check_status_label = ctk.CTkLabel(
            check_controls,
            text="Scan face or QR code to CHECK IN/OUT",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="blue"
        )
        self.check_status_label.pack(side="right", padx=20, pady=15)
    
    def create_status_area(self):
        """Create status and message area"""
        self.status_frame = ctk.CTkFrame(self.main_frame, height=100)
        self.status_frame.pack(fill="x", padx=30, pady=20)
        self.status_frame.pack_propagate(False)
        
        # Status indicator
        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="● READY - Please position yourself in front of the camera",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="green"
        )
        self.status_label.pack(pady=15)
        
        # Help information
        help_text = "⌨️ SHORTCUTS: 1=Clock Mode • 2=Check Mode • Enter=Manual Entry • +=Activate Camera • *=Export • -=Exit"
        help_label = ctk.CTkLabel(
            self.status_frame,
            text=help_text,
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        help_label.pack(pady=5)
        
        # Message area (initially hidden)
        self.message_frame = ctk.CTkFrame(self.main_frame)
        self.message_label = ctk.CTkLabel(
            self.message_frame,
            text="",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        self.message_label.pack(pady=20)
    
    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Original shortcuts (for full keyboards)
        self.root.bind('<F1>', lambda e: self.show_help())                 # F1 for help
        self.root.bind('<F4>', lambda e: self.toggle_kiosk_mode())          # F4 to toggle fullscreen
        self.root.bind('<Escape>', lambda e: self.clear_message())         # ESC to clear messages
        self.root.bind('<Alt-F4>', lambda e: self.quit_app())
        
        # Numpad shortcuts (for numpad-only setups)
        self.root.bind('<KP_Subtract>', lambda e: self.quit_app())          # Numpad -
        self.root.bind('<minus>', lambda e: self.quit_app())                # Regular -
        self.root.bind('<KP_Add>', lambda e: self.toggle_camera())          # Numpad + (toggle camera)
        self.root.bind('<plus>', lambda e: self.toggle_camera())            # Regular + (toggle camera)
        self.root.bind('<KP_Multiply>', lambda e: self.export_daily_csv())  # Numpad * (export CSV)
        self.root.bind('<asterisk>', lambda e: self.export_daily_csv())     # Regular * (export CSV)
        
        # Additional convenient shortcuts
        self.root.bind('<Return>', lambda e: self.show_manual_entry())      # Enter key
        self.root.bind('<KP_Enter>', lambda e: self.show_manual_entry())    # Numpad Enter
        
        # Attendance mode switching
        self.root.bind('<Key-1>', lambda e: self.set_attendance_mode("CLOCK"))  # 1 for CLOCK mode
        self.root.bind('<KP_1>', lambda e: self.set_attendance_mode("CLOCK"))   # Numpad 1 for CLOCK mode
        self.root.bind('<Key-2>', lambda e: self.set_attendance_mode("CHECK"))  # 2 for CHECK mode
        self.root.bind('<KP_2>', lambda e: self.set_attendance_mode("CHECK"))   # Numpad 2 for CHECK mode
        
        # Employee registration
        self.root.bind('<Key-r>', lambda e: self.show_employee_registration_dialog())  # R for registration
        self.root.bind('<Key-R>', lambda e: self.show_employee_registration_dialog())  # R for registration
        
        # Focus on root to capture keys - use multiple methods for reliability
        self.root.focus_set()
        self.root.focus_force()
        
        # Make the window always stay on top and grab focus
        self.root.attributes('-topmost', True)
        self.root.lift()
        
        print("[KEYBOARD DEBUG] Keyboard shortcuts configured and focus set")
    
    def toggle_kiosk_mode(self):
        """Toggle between fullscreen and windowed mode (for debugging)"""
        try:
            if self.root.attributes('-fullscreen'):
                # Switch to windowed mode
                self.root.attributes('-fullscreen', False)
                self.root.overrideredirect(False)
                self.root.geometry("1200x800+100+100")
                print("[KIOSK DEBUG] Switched to windowed mode")
            else:
                # Switch back to fullscreen
                self.setup_kiosk_mode()
                print("[KIOSK DEBUG] Switched back to fullscreen mode")
        except Exception as e:
            print(f"[KIOSK DEBUG] Toggle failed: {e}")
    
    def show_help(self):
        """Show help message"""
        help_text = ("KIOSK ATTENDANCE SYSTEM\n\n" +
                    "MANUAL RECOGNITION MODE:\n" +
                    "● Press + (Plus) to activate camera for face recognition\n" +
                    "● Or enter your Employee ID and press ENTER\n" +
                    "● Or scan QR/Barcode with Employee ID\n\n" +
                    "WORKFLOW:\n" +
                    "1. Press + to turn ON camera\n" +
                    "2. Position your face for recognition\n" +
                    "3. Confirm recognition in dialog\n" +
                    "4. Camera turns OFF automatically\n" +
                    "5. If wrong recognition, camera reactivates for retry\n\n" +
                    "CONTROLS:\n" +
                    "• + = Activate Camera for Recognition\n" +
                    "• Enter = Manual Entry\n" +
                    "• R = Register New Employee\n" +
                    "• * = Export Today's Records (with confirmation)\n" +
                    "• F1 = Show Help\n" +
                    "• F4 = Toggle Fullscreen\n" +
                    "• - = Exit (with confirmation)\n" +
                    "• ESC = Clear Messages")
        
        self.message_frame.pack(fill="x", padx=30, pady=10)
        self.message_label.configure(text=help_text, text_color="cyan")
        self.status_label.configure(text="● HELP DISPLAYED", text_color="cyan")
        
        # Auto-clear after 10 seconds
        self.root.after(10000, self.clear_message)
    
    def show_employee_registration_dialog(self):
        """Show employee registration dialog with camera preview"""
        try:
            # *** CRITICAL: Stop main camera COMPLETELY when registration button is pressed ***
            print("[REG DEBUG] Registration button pressed - stopping main camera completely")
            
            # Stop the main camera and update display
            if self.camera_active:
                self.stop_camera()
                print("[REG DEBUG] Main camera stopped for registration")
            
            self.main_camera_paused = True  # Set flag to prevent camera restart during registration
            
            # Also stop face recognition background processing temporarily
            if hasattr(self, 'face_recognition') and self.face_recognition:
                print("[REG DEBUG] Pausing background face recognition during registration")
            
            # Create registration dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Employee Registration")
            dialog.geometry("950x650")  # Even larger to accommodate bigger camera preview
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.resizable(False, False)
            
            # Center dialog on screen
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (950 // 2)
            y = (dialog.winfo_screenheight() // 2) - (650 // 2)
            dialog.geometry(f"950x650+{x}+{y}")
            
            dialog.configure(bg='#f0f8ff')
            
            # Ensure dialog closes properly and resumes camera
            def on_dialog_close():
                print("[REG DEBUG] Dialog close event triggered")
                self.close_registration_dialog(dialog)
            
            dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
            
            # Title
            title_label = tk.Label(
                dialog,
                text="👤 REGISTER NEW EMPLOYEE",
                font=("Arial", 18, "bold"),
                bg='#f0f8ff',
                fg='#2E86C1'
            )
            title_label.pack(pady=10)
            
            # Main container frame
            main_container = tk.Frame(dialog, bg='#f0f8ff')
            main_container.pack(fill="both", expand=True, padx=20, pady=10)
            
            # Left side - Form
            form_frame = tk.Frame(main_container, bg='#f0f8ff', width=350)
            form_frame.pack(side="left", fill="y", padx=10)
            form_frame.pack_propagate(False)
            
            # Right side - Camera preview (larger to accommodate bigger camera feed)
            camera_frame = tk.Frame(main_container, bg='#f0f8ff', width=520, height=450)
            camera_frame.pack(side="right", fill="y", padx=10)
            camera_frame.pack_propagate(False)
            
            # Form title
            form_title = tk.Label(
                form_frame,
                text="📝 EMPLOYEE DETAILS",
                font=("Arial", 14, "bold"),
                bg='#f0f8ff',
                fg='#2E86C1'
            )
            form_title.grid(row=0, column=0, columnspan=2, pady=(10, 20))
            
            # Employee ID
            tk.Label(form_frame, text="Employee ID:", font=("Arial", 12, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=1, column=0, sticky="e", padx=10, pady=10)
            self.reg_employee_id = tk.Entry(form_frame, font=("Arial", 12), width=18)
            self.reg_employee_id.grid(row=1, column=1, padx=10, pady=10)
            
            # Employee Name
            tk.Label(form_frame, text="Full Name:", font=("Arial", 12, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=2, column=0, sticky="e", padx=10, pady=10)
            self.reg_employee_name = tk.Entry(form_frame, font=("Arial", 12), width=18)
            self.reg_employee_name.grid(row=2, column=1, padx=10, pady=10)
            
            # Department
            tk.Label(form_frame, text="Department:", font=("Arial", 12, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=3, column=0, sticky="e", padx=10, pady=10)
            self.reg_department = tk.Entry(form_frame, font=("Arial", 12), width=18)
            self.reg_department.grid(row=3, column=1, padx=10, pady=10)
            
            # Role Selection
            tk.Label(form_frame, text="Role:", font=("Arial", 12, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=4, column=0, sticky="e", padx=10, pady=10)
            
            # Create role dropdown
            self.reg_role_var = tk.StringVar(value="Staff")
            role_frame = tk.Frame(form_frame, bg='#f0f8ff')
            role_frame.grid(row=4, column=1, padx=10, pady=10, sticky="w")
            
            role_options = ["Staff", "Security"]
            self.reg_role_menu = tk.OptionMenu(role_frame, self.reg_role_var, *role_options)
            self.reg_role_menu.config(
                font=("Arial", 11),
                bg='white',
                fg='#2E86C1',
                width=15,
                borderwidth=1,
                relief="solid"
            )
            self.reg_role_menu.pack(side="left")
            
            
            # Status label
            self.reg_status_label = tk.Label(
                form_frame,
                text="Fill in the details above, then click 'Start Face Capture'",
                font=("Arial", 10),
                bg='#f0f8ff',
                fg='#666666',
                wraplength=320
            )
            self.reg_status_label.grid(row=5, column=0, columnspan=2, pady=15)
            
            # Camera preview setup
            camera_title = tk.Label(
                camera_frame,
                text="📷 CAMERA PREVIEW",
                font=("Arial", 14, "bold"),
                bg='#f0f8ff',
                fg='#2E86C1'
            )
            camera_title.pack(pady=(10, 5))
            
            # Camera display label - no fixed size to allow image to control dimensions
            self.reg_camera_label = tk.Label(
                camera_frame,
                text="Camera will start when capture begins",
                font=("Arial", 12),
                bg='#333333',
                fg='white'
            )
            self.reg_camera_label.pack(pady=10)
            
            # Buttons frame (at bottom)
            button_frame = tk.Frame(dialog, bg='#f0f8ff')
            button_frame.pack(side="bottom", pady=20)
            
            # Start capture button
            self.capture_btn = tk.Button(
                button_frame,
                text="START FACE CAPTURE",
                font=("Arial", 14, "bold"),
                bg='#28a745',
                fg='white',
                width=20,
                height=2,
                command=lambda: self.start_face_capture_with_preview(dialog)
            )
            self.capture_btn.pack(side="left", padx=10)
            
            # Cancel button
            cancel_btn = tk.Button(
                button_frame,
                text="CANCEL",
                font=("Arial", 12),
                bg='#dc3545',
                fg='white',
                width=20,
                height=2,
                command=dialog.destroy
            )
            cancel_btn.pack(side="left", padx=10)
            
            # Bind keyboard shortcuts
            dialog.bind('<Return>', lambda e: self.start_face_capture_with_preview(dialog))
            dialog.bind('<Escape>', lambda e: self.close_registration_dialog(dialog))
            
            # Focus on employee ID field
            self.reg_employee_id.focus_set()
            
            # Store dialog reference and initialize variables
            self.registration_dialog = dialog
            self.reg_camera_active = False
            self.reg_capture_active = False
            
        except Exception as e:
            print(f"[REG DEBUG] Error showing registration dialog: {e}")
            # Resume main camera if dialog creation failed
            self.main_camera_paused = False
            print("[REG DEBUG] Main camera resumed after dialog creation error")
            self.show_error_message(f"❌ Registration error: {str(e)}")
    
    def close_registration_dialog(self, dialog):
        """Close registration dialog and resume main camera"""
        try:
            print("[REG DEBUG] Closing registration dialog and resuming main camera")
            
            # Stop registration camera activities
            self.reg_camera_active = False
            self.reg_capture_active = False
            
            # Stop the registration camera
            if hasattr(self, 'camera_manager') and self.camera_manager:
                self.camera_manager.stop_camera()
                print("[REG DEBUG] Registration camera stopped")
            
            # Resume main camera processing IMMEDIATELY
            if hasattr(self, 'main_camera_paused'):
                self.main_camera_paused = False
                print("[REG DEBUG] Main camera flag cleared")
            
            # Restart the main camera
            if not self.camera_active:
                print("[REG DEBUG] Restarting main camera after registration")
                self.start_camera()
            
            # Resume background face recognition
            if hasattr(self, 'face_recognition') and self.face_recognition:
                print("[REG DEBUG] Background face recognition resumed")
            
            # Force update of camera display
            if hasattr(self, 'camera_label'):
                self.camera_label.configure(text="📷 Camera Resuming...")
                print("[REG DEBUG] Camera label updated")
            
            dialog.destroy()
            
            # Force focus back to main window
            self.root.focus_force()
            print("[REG DEBUG] Focus returned to main window")
            
        except Exception as e:
            print(f"[REG DEBUG] Error closing dialog: {e}")
            # Ensure main camera is resumed even if there's an error
            if hasattr(self, 'main_camera_paused'):
                self.main_camera_paused = False
                print("[REG DEBUG] Main camera flag force-cleared after error")
            
            # Restart main camera if not active
            if not self.camera_active:
                print("[REG DEBUG] Force-restarting main camera after error")
                self.start_camera()
    
    def start_face_capture_with_preview(self, dialog):
        """Start face capture with live camera preview and bounding boxes"""
        try:
            print("[REG DEBUG] Start face capture button pressed")
            
            # Validate form inputs
            employee_id = self.reg_employee_id.get().strip()
            employee_name = self.reg_employee_name.get().strip()
            department = self.reg_department.get().strip()
            role = self.reg_role_var.get()  # Get selected role
            
            if not employee_id or not employee_name or not department:
                self.reg_status_label.configure(
                    text="❌ Please fill in all fields",
                    fg='red'
                )
                return
            
            # Check if employee ID already exists
            if self.db.get_employee(employee_id):
                self.reg_status_label.configure(
                    text="❌ Employee ID already exists",
                    fg='red'
                )
                return
            
            # Check if camera manager is available (we'll start our own registration camera)
            if not hasattr(self, 'camera_manager') or not self.camera_manager:
                self.reg_status_label.configure(
                    text="❌ Camera manager not available.",
                    fg='red'
                )
                return
            
            # Start registration camera (independent of main camera)
            print("[REG DEBUG] Starting dedicated registration camera")
            if not self.camera_manager.start_camera():
                self.reg_status_label.configure(
                    text="❌ Cannot start registration camera. Please check camera connection.",
                    fg='red'
                )
                return
            
            print("[REG DEBUG] All validations passed, starting capture process")
            
            # Update status and start preview (main camera already paused when dialog opened)
            self.reg_status_label.configure(
                text="📷 Starting camera preview... Position yourself in the frame",
                fg='blue'
            )
            
            # Disable form inputs during capture
            self.reg_employee_id.configure(state='disabled')
            self.reg_employee_name.configure(state='disabled')
            self.reg_department.configure(state='disabled')
            self.reg_role_menu.configure(state='disabled')
            self.capture_btn.configure(state='disabled', text="📷 CAPTURING...")
            
            # Start camera preview for registration
            self.reg_camera_active = True
            self.start_registration_camera_preview()
            
            # Start actual capture process after a short delay
            dialog.after(2000, lambda: self.start_face_capture_process(employee_id, employee_name, department, role, dialog))
            
        except Exception as e:
            print(f"[REG DEBUG] Error starting face capture: {e}")
            import traceback
            traceback.print_exc()
            
            # Re-enable form on error
            try:
                self.reg_employee_id.configure(state='normal')
                self.reg_employee_name.configure(state='normal')
                self.reg_department.configure(state='normal')
                self.reg_role_menu.configure(state='normal')
                self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
            except:
                pass
                
            self.reg_status_label.configure(
                text=f"❌ Capture error: {str(e)}",
                fg='red'
            )
    
    def start_registration_camera_preview(self):
        """Start camera preview with face detection bounding boxes"""
        try:
            if not self.reg_camera_active:
                return
                
            if hasattr(self, 'camera_manager') and self.camera_manager and self.camera_manager.cap is not None:
                # Check if camera is still open
                if not self.camera_manager.cap.isOpened():
                    print("[REG DEBUG] Camera connection lost during preview")
                    self.reg_camera_active = False
                    return
                
                ret, frame = self.camera_manager.cap.read()
                
                if ret and frame is not None:
                    # Create a copy for drawing
                    display_frame = frame.copy()
                    height, width = frame.shape[:2]
                    display_width, display_height = 480, 360
                    
                    # Use ultra light detection for fast real-time feedback
                    detected_faces = []
                    if hasattr(self, 'ultra_light_detector') and self.ultra_light_detector:
                        try:
                            detected_faces = self.ultra_light_detector.detect_faces_for_attendance(frame)
                            print(f"[REG ULTRA] Detected {len(detected_faces) if detected_faces else 0} faces in registration")
                        except Exception as ultra_error:
                            print(f"[REG ULTRA] Ultra light detection error: {ultra_error}")
                    
                    # Draw face detection results
                    if detected_faces:
                        # Get the best face for registration
                        best_face = self.ultra_light_detector.get_best_face(detected_faces)
                        
                        for face_data in detected_faces:
                            x1, y1, x2, y2 = face_data['bbox']
                            confidence = face_data['confidence']
                            
                            # Scale coordinates to display frame
                            display_x1 = int(x1 * display_width / width)
                            display_y1 = int(y1 * display_height / height)
                            display_x2 = int(x2 * display_width / width)
                            display_y2 = int(y2 * display_height / height)
                            
                            # Determine if this is the best face
                            is_best = best_face and face_data['id'] == best_face['id']
                            
                            # Color coding - green for good detection, yellow for others
                            if confidence > 0.7 and is_best:
                                color = (0, 255, 0)  # Green for good quality
                                label = f"GOOD QUALITY ({confidence:.2f})"
                                thickness = 3
                            elif is_best:
                                color = (0, 255, 255)  # Yellow for acceptable
                                label = f"FACE DETECTED ({confidence:.2f})"
                                thickness = 2
                            else:
                                color = (255, 255, 0)  # Light blue for other faces
                                label = f"Other Face ({confidence:.2f})"
                                thickness = 1
                            
                            # Draw bounding box
                            cv2.rectangle(display_frame, (display_x1, display_y1), (display_x2, display_y2), color, thickness)
                            
                            # Draw label with background
                            label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 1)[0]
                            cv2.rectangle(display_frame, 
                                         (display_x1, display_y1 - label_size[1] - 10), 
                                         (display_x1 + label_size[0], display_y1), 
                                         color, -1)
                            cv2.putText(display_frame, label, (display_x1, display_y1 - 5), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 1)
                        
                        # Add instruction text at bottom
                        cv2.putText(display_frame, "FACE DETECTED - READY FOR CAPTURE", (30, display_height - 20), 
                                  cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                    else:
                        # No faces detected - add instruction text
                        cv2.putText(display_frame, "POSITION FACE IN FRAME", (30, 40), 
                                  cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                        cv2.putText(display_frame, "Look directly at the camera", (30, display_height - 20), 
                                  cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 1)
                    
                    # Add FPS indicator
                    if hasattr(self, 'ultra_light_detector') and self.ultra_light_detector:
                        fps = self.ultra_light_detector.get_performance_stats()['fps']
                        cv2.putText(display_frame, f"FPS: {fps:.1f}", (display_width - 100, 30), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                    
                    try:
                        # Resize and convert for display - larger size for better visibility
                        display_frame = cv2.resize(display_frame, (display_width, display_height))
                        display_frame = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
                        
                        # Convert to PhotoImage
                        from PIL import Image, ImageTk
                        image = Image.fromarray(display_frame)
                        photo = ImageTk.PhotoImage(image)
                        
                        # Update camera display
                        if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                            self.reg_camera_label.configure(image=photo)
                            self.reg_camera_label.image = photo  # Keep a reference
                    except Exception as display_error:
                        print(f"[REG DEBUG] Display error: {display_error}")
                        # Don't crash on display errors
            
            # Schedule next frame update
            if self.reg_camera_active and hasattr(self, 'registration_dialog'):
                try:
                    self.registration_dialog.after(20, self.start_registration_camera_preview)  # 50 FPS for smoother registration
                except:
                    # Dialog might be destroyed
                    self.reg_camera_active = False
                
        except Exception as e:
            print(f"[REG DEBUG] Camera preview error: {e}")
            import traceback
            traceback.print_exc()
            # Continue trying if camera is still active, but with shorter delay for smoother recovery
            if self.reg_camera_active and hasattr(self, 'registration_dialog'):
                try:
                    self.registration_dialog.after(50, self.start_registration_camera_preview)  # 20 FPS fallback
                except:
                    self.reg_camera_active = False
    
    def start_face_capture_process(self, employee_id, employee_name, department, role, dialog):
        """Start the actual face capture process"""
        try:
            print(f"[REG DEBUG] Starting face capture process for {employee_id} ({role})")
            
            # Validate that we still have camera access
            if not hasattr(self, 'camera_manager') or not self.camera_manager or not self.camera_manager.cap:
                self.reg_status_label.configure(
                    text="❌ Camera lost during process. Please try again.",
                    fg='red'
                )
                return
            
            self.reg_capture_active = True
            
            # Update status
            self.reg_status_label.configure(
                text="🎯 Capturing face vectors... Hold still!",
                fg='green'
            )
            
            # Start capture in separate thread with better error handling
            try:
                capture_thread = threading.Thread(
                    target=self.capture_face_vectors_with_preview,
                    args=(employee_id, employee_name, department, role, dialog),
                    daemon=True
                )
                capture_thread.start()
                print("[REG DEBUG] Capture thread started successfully")
                
            except Exception as thread_error:
                print(f"[REG DEBUG] Error starting capture thread: {thread_error}")
                self.reg_capture_active = False
                self.reg_status_label.configure(
                    text=f"❌ Threading error: {str(thread_error)}",
                    fg='red'
                )
            
        except Exception as e:
            print(f"[REG DEBUG] Error starting capture process: {e}")
            import traceback
            traceback.print_exc()
            
            self.reg_capture_active = False
            self.reg_status_label.configure(
                text=f"❌ Capture failed: {str(e)}",
                fg='red'
            )
    
    def start_face_capture(self, dialog):
        """Start face capture process for employee registration"""
        try:
            # Validate form inputs
            employee_id = self.reg_employee_id.get().strip()
            employee_name = self.reg_employee_name.get().strip()
            department = self.reg_department.get().strip()
            
            if not employee_id or not employee_name or not department:
                self.reg_status_label.configure(
                    text="❌ Please fill in all fields",
                    fg='red'
                )
                return
            
            # Check if employee ID already exists
            if self.db.get_employee(employee_id):
                self.reg_status_label.configure(
                    text="❌ Employee ID already exists",
                    fg='red'
                )
                return
            
            # Check if camera is available
            if not self.camera_active:
                self.reg_status_label.configure(
                    text="❌ Camera is not active. Please enable camera first.",
                    fg='red'
                )
                return
            
            # Update status
            self.reg_status_label.configure(
                text="📷 Starting face capture... Look at the camera!",
                fg='blue'
            )
            
            # Disable form inputs during capture
            self.reg_employee_id.configure(state='disabled')
            self.reg_employee_name.configure(state='disabled')
            self.reg_department.configure(state='disabled')
            self.capture_btn.configure(state='disabled', text="📷 CAPTURING...")
            
            # Start face capture in separate thread
            capture_thread = threading.Thread(
                target=self.capture_face_vectors,
                args=(employee_id, employee_name, department, dialog),
                daemon=True
            )
            capture_thread.start()
            
        except Exception as e:
            print(f"[REG DEBUG] Error starting face capture: {e}")
            self.reg_status_label.configure(
                text=f"❌ Capture error: {str(e)}",
                fg='red'
            )
    
    def capture_face_vectors(self, employee_id, employee_name, department, dialog):
        """Capture multiple face images and create face vectors with visual feedback"""
        try:
            print(f"[REG DEBUG] Starting enhanced face capture for {employee_id}")
            
            captured_frames = []  # Store frames for robust embedding creation
            captured_vectors = []
            capture_count = 0
            max_captures = 15  # Increased to capture more frames for better quality
            capture_duration = 5  # 5 seconds total capture time
            min_face_size = 60  # Minimum face size for quality
            
            start_time = time.time()
            
            # Temporarily stop the preview to show capture feedback
            original_preview_state = self.reg_camera_active
            self.reg_camera_active = False
            
            while capture_count < max_captures and (time.time() - start_time) < capture_duration:
                if hasattr(self, 'camera_manager') and self.camera_manager.cap is not None:
                    ret, frame = self.camera_manager.cap.read()
                    
                    if ret and frame is not None:
                        # Create display frame for visual feedback
                        display_frame = frame.copy()
                        height, width = frame.shape[:2]
                        display_width, display_height = 480, 360
                        
                        # Try to detect and validate face quality using ultra light detection
                        face_quality_good = False
                        best_face_data = None
                        
                        if hasattr(self, 'ultra_light_detector') and self.ultra_light_detector:
                            try:
                                detected_faces = self.ultra_light_detector.detect_faces_for_attendance(frame)
                                if detected_faces:
                                    face_detected = True
                                    best_face = self.ultra_light_detector.get_best_face(detected_faces)
                                    
                                    if best_face:
                                        best_face_data = best_face
                                        x1, y1, x2, y2 = best_face['bbox']
                                        face_width = x2 - x1
                                        face_height = y2 - y1
                                        confidence = best_face['confidence']
                                        
                                        # Quality checks
                                        if (face_width >= min_face_size and 
                                            face_height >= min_face_size and 
                                            confidence >= 0.8 and
                                            face_width < width * 0.8 and  # Not too large (too close)
                                            face_height < height * 0.8):
                                            face_quality_good = True
                                    
                                    # Draw detection box
                                    for face_data in detected_faces:
                                        x1, y1, x2, y2 = face_data['bbox']
                                        confidence = face_data['confidence']
                                        is_best = best_face and face_data['id'] == best_face['id']
                                        
                                        # Scale coordinates to display frame
                                        display_x1 = int(x1 * display_width / width)
                                        display_y1 = int(y1 * display_height / height)
                                        display_x2 = int(x2 * display_width / width)
                                        display_y2 = int(y2 * display_height / height)
                                        
                                        # Color coding for capture feedback
                                        if face_quality_good and is_best:
                                            color = (0, 255, 0)  # Green for high quality
                                            thickness = 3
                                        elif is_best:
                                            color = (0, 255, 255)  # Yellow for medium quality
                                            thickness = 2
                                        else:
                                            color = (0, 165, 255)  # Orange for detected faces
                                            thickness = 1
                                        
                                        # Draw bounding box
                                        cv2.rectangle(display_frame, (display_x1, display_y1), (display_x2, display_y2), color, thickness)
                                        
                                        # Show quality and vectorization status
                                        if is_best:
                                            if face_quality_good:
                                                label = f"HIGH QUALITY - CAPTURING {capture_count}/{max_captures}"
                                            else:
                                                label = f"ADJUSTING POSITION..."
                                            label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                                            cv2.rectangle(display_frame, 
                                                         (display_x1, display_y1 - label_size[1] - 15), 
                                                         (display_x1 + label_size[0], display_y1), 
                                                         color, -1)
                                            cv2.putText(display_frame, label, (display_x1, display_y1 - 5), 
                                                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)
                            except Exception as ultra_error:
                                print(f"[REG CAPTURE] Ultra detection error: {ultra_error}")
                        
                        # Capture frame if quality is good
                        if face_quality_good and len(captured_frames) < max_captures:
                            captured_frames.append(frame.copy())
                            capture_count += 1
                            
                            # Show success feedback on display
                            cv2.putText(display_frame, f"✓ FRAME {capture_count} CAPTURED!", (30, display_height - 40), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                            
                            # Update status on UI thread
                            progress_text = f"📸 Capturing frames... {capture_count}/{max_captures} high-quality frames captured"
                            self.root.after(0, lambda: self.reg_status_label.configure(
                                text=progress_text, fg='green'
                            ))
                            
                            print(f"[REG DEBUG] Captured high-quality frame {capture_count}/{max_captures}")
                            
                            # Small delay between captures to get different poses
                            time.sleep(0.2)
                        else:
                            # Provide feedback on why frame wasn't captured
                            if not face_detected:
                                feedback = "NO FACE DETECTED"
                                color = (0, 0, 255)  # Red
                            elif not face_quality_good:
                                feedback = "ADJUSTING FOR BETTER QUALITY..."
                                color = (0, 255, 255)  # Yellow
                            else:
                                feedback = f"ENOUGH FRAMES CAPTURED ({capture_count})"
                                color = (0, 255, 0)  # Green
                                
                            cv2.putText(display_frame, feedback, (30, display_height - 40), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
                        
                        # Add progress bar
                        progress_width = int((capture_count / max_captures) * (display_width - 60))
                        cv2.rectangle(display_frame, (30, display_height - 20), (display_width - 30, display_height - 10), (100, 100, 100), -1)
                        if progress_width > 0:
                            cv2.rectangle(display_frame, (30, display_height - 20), (30 + progress_width, display_height - 10), (0, 255, 0), -1)
                        
                        # Update camera display with capture feedback
                        try:
                            display_frame = cv2.resize(display_frame, (display_width, display_height))
                            display_frame = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
                            
                            from PIL import Image, ImageTk
                            image = Image.fromarray(display_frame)
                            photo = ImageTk.PhotoImage(image)
                            
                            if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                                self.reg_camera_label.configure(image=photo)
                                self.reg_camera_label.image = photo
                        except Exception as display_error:
                            print(f"[REG CAPTURE] Display error: {display_error}")
                        
                        # Small delay between captures
                        time.sleep(0.2)
                
                time.sleep(0.1)  # Small delay to prevent excessive CPU usage
            
            # Restore preview state
            self.reg_camera_active = original_preview_state
            
            # Process captured frames to create robust embedding
            if len(captured_frames) >= 5:  # Need at least 5 good quality frames
                # Update status to show processing
                self.root.after(0, lambda: self.reg_status_label.configure(
                    text=f"🧠 Creating robust face embedding from {len(captured_frames)} frames...",
                    fg='blue'
                ))
                
                # Create robust embedding using the enhanced method
                robust_embedding = self.face_recognition.create_robust_face_embedding(captured_frames)
                
                if robust_embedding is not None:
                    # Update status to show database saving
                    self.root.after(0, lambda: self.reg_status_label.configure(
                        text=f"💾 Saving to database... (robust embedding from {len(captured_frames)} frames)",
                        fg='blue'
                    ))
                    
                    # Save employee to database
                    success = self.db.add_employee_with_face_vector(
                        employee_id=employee_id,
                        name=employee_name,
                        face_vector=robust_embedding.tolist(),
                        department=department
                    )
                    
                    if success:
                        # Update status on UI thread
                        self.root.after(0, lambda: self.reg_status_label.configure(
                            text=f"✅ Employee registered successfully! (ArcFace robust embedding from {len(captured_frames)} frames)",
                            fg='green'
                        ))
                        
                        print(f"[REG DEBUG] Employee {employee_id} registered with robust ArcFace embedding from {len(captured_frames)} frames")
                        
                        # Close dialog after 3 seconds
                        self.root.after(3000, dialog.destroy)
                        
                        # Show success message
                        self.root.after(100, lambda: self.show_success_message(
                            f"👤 Employee {employee_name} registered successfully with high-accuracy ArcFace embedding!"
                        ))
                        
                        # Reload face recognition data
                        self.root.after(500, self.load_known_faces)
                        
                    else:
                        self.root.after(0, lambda: self.reg_status_label.configure(
                            text="❌ Database error - failed to save employee",
                            fg='red'
                        ))
                else:
                    self.root.after(0, lambda: self.reg_status_label.configure(
                        text="❌ Failed to create face embedding - please try again",
                        fg='red'
                    ))
            else:
                # Not enough high-quality frames captured
                self.root.after(0, lambda: self.reg_status_label.configure(
                    text=f"❌ Could not capture enough high-quality frames ({len(captured_frames)}/5 minimum). Please try again.",
                    fg='red'
                ))
            
            # Re-enable form on UI thread
            self.root.after(0, lambda: [
                self.reg_employee_id.configure(state='normal'),
                self.reg_employee_name.configure(state='normal'),
                self.reg_department.configure(state='normal'),
                self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
            ])
                
        except Exception as e:
            print(f"[REG DEBUG] Error during face capture: {e}")
            import traceback
            traceback.print_exc()
            
            # Restore preview state
            if 'original_preview_state' in locals():
                self.reg_camera_active = original_preview_state
            
            # Show error on UI thread
            self.root.after(0, lambda: self.reg_status_label.configure(
                text=f"❌ Capture failed: {str(e)}",
                fg='red'
            ))
            
            # Re-enable form on UI thread
            self.root.after(0, lambda: [
                self.reg_employee_id.configure(state='normal'),
                self.reg_employee_name.configure(state='normal'),
                self.reg_department.configure(state='normal'),
                self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
            ])
                
        except Exception as e:
            print(f"[REG DEBUG] Error during face capture: {e}")
            import traceback
            traceback.print_exc()
            
            # Restore preview state
            if 'original_preview_state' in locals():
                self.reg_camera_active = original_preview_state
            
            # Show error on UI thread
            self.root.after(0, lambda: self.reg_status_label.configure(
                text=f"❌ Capture failed: {str(e)}",
                fg='red'
            ))
            
            # Re-enable form on UI thread
            self.root.after(0, lambda: [
                self.reg_employee_id.configure(state='normal'),
                self.reg_employee_name.configure(state='normal'),
                self.reg_department.configure(state='normal'),
                self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
            ])
                
        except Exception as e:
            print(f"[REG DEBUG] Error during face capture: {e}")
            import traceback
            traceback.print_exc()
            
            # Show error on UI thread
            self.root.after(0, lambda: self.reg_status_label.configure(
                text=f"❌ Capture failed: {str(e)}",
                fg='red'
            ))
            
            # Re-enable form on UI thread
            self.root.after(0, lambda: [
                self.reg_employee_id.configure(state='normal'),
                self.reg_employee_name.configure(state='normal'),
                self.reg_department.configure(state='normal'),
                self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
            ])
    
    def capture_face_vectors_with_preview(self, employee_id, employee_name, department, role, dialog):
        """Improved face capture with guided instructions for better accuracy"""
        try:
            import numpy as np
            print(f"[IMPROVED REG] Starting guided face capture for {employee_id} ({role})")
            
            # Pause background processing
            print("[IMPROVED REG] Pausing background face processing during registration")
            self.face_recognition.stop_background_processing()
            
            # Define guided instructions for diverse face capture
            capture_instructions = [
                "Look straight at the camera with a NEUTRAL expression",
                "Turn your head slightly to the LEFT (keep looking at camera)",
                "Turn your head slightly to the RIGHT (keep looking at camera)",
                "Look straight and SMILE naturally",
                "Look straight with a SERIOUS expression",
                "Tilt your head slightly UP (chin up a bit)",
                "Tilt your head slightly DOWN (chin down a bit)",
                "Look straight again - FINAL CAPTURE"
            ]
            
            captured_vectors = []
            total_instructions = len(capture_instructions)
            
            # Process each instruction
            for instruction_idx, instruction in enumerate(capture_instructions):
                if not self.reg_capture_active or not self.reg_camera_active:
                    break
                
                progress = f"{instruction_idx + 1}/{total_instructions}"
                
                # Update UI with current instruction
                def update_instruction_ui(progress=progress, instruction=instruction):
                    try:
                        if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                            self.reg_status_label.configure(
                                text=f"📋 {progress}: {instruction}",
                                fg='blue'
                            )
                        if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                            self.reg_camera_label.configure(
                                text=f"📋 INSTRUCTION {progress}\n\n{instruction}\n\nGet ready...",
                                bg='#007bff',
                                fg='white'
                            )
                    except:
                        pass
                
                self.root.after(0, update_instruction_ui)
                
                # Give user time to read and position (3 seconds)
                time.sleep(3.0)
                
                # Update UI to show capturing
                def update_capturing_ui(progress=progress):
                    try:
                        if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                            self.reg_status_label.configure(
                                text=f"📸 {progress}: Capturing... Hold still!",
                                fg='orange'
                            )
                        if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                            self.reg_camera_label.configure(
                                text=f"📸 CAPTURING {progress}\n\nHOLD STILL!\n\nCapturing face...",
                                bg='#fd7e14',
                                fg='white'
                            )
                    except:
                        pass
                
                self.root.after(0, update_capturing_ui)
                
                # Attempt to capture face vector for this pose
                success = self._capture_single_pose_improved(instruction)
                
                if success:
                    captured_vectors.append(success)
                    
                    # Update success UI
                    def update_success_ui(progress=progress):
                        try:
                            if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                                self.reg_status_label.configure(
                                    text=f"✅ {progress}: Captured! Moving to next pose...",
                                    fg='green'
                                )
                        except:
                            pass
                    
                    self.root.after(0, update_success_ui)
                    time.sleep(1.5)  # Brief pause before next instruction
                    
                else:
                    # Retry once
                    def update_retry_ui(progress=progress):
                        try:
                            if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                                self.reg_status_label.configure(
                                    text=f"⚠️ {progress}: Retrying capture...",
                                    fg='orange'
                                )
                        except:
                            pass
                    
                    self.root.after(0, update_retry_ui)
                    time.sleep(1.0)
                    
                    success = self._capture_single_pose_improved(instruction)
                    if success:
                        captured_vectors.append(success)
                    else:
                        print(f"[IMPROVED REG] Skipping pose: {instruction}")
            
            # Stop camera preview
            self.reg_camera_active = False
            self.reg_capture_active = False
            
            print(f"[IMPROVED REG] Capture completed. Total vectors: {len(captured_vectors)}")
            
            # Process results with weighted averaging
            if len(captured_vectors) >= 3:  # Need at least 3 good captures
                def update_processing_status():
                    try:
                        if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                            self.reg_status_label.configure(text="🔄 Processing diverse face data...", fg='blue')
                    except:
                        pass
                
                self.root.after(0, update_processing_status)
                
                # Create weighted average - give more weight to front-facing captures
                vectors = []
                weights = []
                
                for item in captured_vectors:
                    vectors.append(item['vector'])
                    instruction = item['instruction'].lower()
                    if 'straight' in instruction or 'neutral' in instruction:
                        weights.append(1.5)  # Higher weight for straight/neutral views
                    elif 'smile' in instruction or 'serious' in instruction:
                        weights.append(1.3)  # Medium weight for expression changes
                    else:
                        weights.append(1.0)  # Normal weight for pose changes
                
                # Calculate weighted average
                weights = np.array(weights)
                weights = weights / weights.sum()  # Normalize weights
                
                weighted_avg = np.zeros_like(vectors[0])
                for i, vector in enumerate(vectors):
                    weighted_avg += vector * weights[i]
                
                # Save employee to database
                try:
                    success = self.db.add_employee_with_face_vector(
                        employee_id=employee_id,
                        name=employee_name,
                        face_vector=weighted_avg.tolist(),
                        department=department,
                        role=role
                    )
                    
                    if success:
                        print(f"[IMPROVED REG] Employee {employee_id} ({role}) registered successfully with {len(captured_vectors)} diverse face vectors")
                        
                        # Update success status
                        vector_count = len(captured_vectors)
                        def update_success_status():
                            try:
                                if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                                    self.reg_status_label.configure(
                                        text=f"✅ Registration successful! ({vector_count} diverse vectors captured)",
                                        fg='green'
                                    )
                                if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                                    self.reg_camera_label.configure(
                                        text=f"✅ REGISTRATION COMPLETE!\n\n{employee_name}\nID: {employee_id}\nDiverse Vectors: {vector_count}",
                                        bg='#28a745',
                                        fg='white'
                                    )
                            except Exception as ui_error:
                                print(f"[IMPROVED REG] UI update error: {ui_error}")
                        
                        self.root.after(0, update_success_status)
                        
                        # Close dialog after 3 seconds
                        def close_dialog():
                            try:
                                if dialog.winfo_exists():
                                    dialog.destroy()
                            except:
                                pass
                        
                        self.root.after(3000, close_dialog)
                        
                        # Show success message in main app
                        def show_main_success():
                            try:
                                self.show_success_message(
                                    f"👤 Employee {employee_name} registered with improved face recognition!"
                                )
                            except Exception as msg_error:
                                print(f"[IMPROVED REG] Success message error: {msg_error}")
                        
                        self.root.after(100, show_main_success)
                        
                        # Reload face recognition data
                        def reload_faces():
                            try:
                                self.load_known_faces()
                            except Exception as reload_error:
                                print(f"[IMPROVED REG] Face reload error: {reload_error}")
                        
                        self.root.after(500, reload_faces)
                        
                    else:
                        print(f"[IMPROVED REG] Database error - failed to save employee {employee_id}")
                        def update_db_error():
                            try:
                                if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                                    self.reg_status_label.configure(text="❌ Database error - failed to save employee", fg='red')
                            except:
                                pass
                        
                        self.root.after(0, update_db_error)
                        
                except Exception as db_error:
                    print(f"[IMPROVED REG] Database operation error: {db_error}")
                    def update_db_error():
                        try:
                            if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                                self.reg_status_label.configure(text=f"❌ Database error: {str(db_error)}", fg='red')
                        except:
                            pass
                    
                    self.root.after(0, update_db_error)
                        
            else:
                # Not enough diverse faces captured
                vector_count = len(captured_vectors)
                print(f"[IMPROVED REG] Insufficient diverse vectors captured: {vector_count}/3 minimum")
                
                def update_insufficient_status():
                    try:
                        if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                            self.reg_status_label.configure(
                                text=f"❌ Insufficient diverse face data ({vector_count}/3 minimum). Please try again.",
                                fg='red'
                            )
                        if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                            self.reg_camera_label.configure(
                                text=f"❌ CAPTURE FAILED\nOnly {vector_count} diverse poses\nMinimum required: 3",
                                bg='#dc3545',
                                fg='white'
                            )
                    except:
                        pass
                
                self.root.after(0, update_insufficient_status)
            
            # Re-enable form on UI thread
            def re_enable_form():
                try:
                    if hasattr(self, 'reg_employee_id') and self.reg_employee_id.winfo_exists():
                        self.reg_employee_id.configure(state='normal')
                    if hasattr(self, 'reg_employee_name') and self.reg_employee_name.winfo_exists():
                        self.reg_employee_name.configure(state='normal')
                    if hasattr(self, 'reg_department') and self.reg_department.winfo_exists():
                        self.reg_department.configure(state='normal')
                    if hasattr(self, 'capture_btn') and self.capture_btn.winfo_exists():
                        self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
                except Exception as form_error:
                    print(f"[IMPROVED REG] Form re-enable error: {form_error}")
            
            # Resume main camera processing
            def resume_main_camera():
                try:
                    self.main_camera_paused = False
                    print("[IMPROVED REG] Main camera flag cleared after capture completion")
                    
                    # Restart main camera if not active
                    if not self.camera_active:
                        print("[IMPROVED REG] Restarting main camera after successful registration")
                        self.start_camera()
                    
                    # Resume background face processing
                    self.face_recognition.start_background_processing()
                    print("[IMPROVED REG] Background face processing resumed")
                except Exception as camera_error:
                    print(f"[IMPROVED REG] Error resuming main camera: {camera_error}")
            
            self.root.after(0, re_enable_form)
            self.root.after(0, resume_main_camera)
                
        except Exception as e:
            print(f"[REG DEBUG] Error during face capture with preview: {e}")
            import traceback
            traceback.print_exc()
            
            # Stop camera
            self.reg_camera_active = False
            self.reg_capture_active = False
            
            # Show error on UI thread with safe handling
            def update_error_status():
                try:
                    if hasattr(self, 'reg_status_label') and self.reg_status_label.winfo_exists():
                        self.reg_status_label.configure(text=f"❌ Capture failed: {str(e)}", fg='red')
                    if hasattr(self, 'reg_camera_label') and self.reg_camera_label.winfo_exists():
                        self.reg_camera_label.configure(text=f"❌ ERROR\n{str(e)}", bg='#dc3545', fg='white')
                except Exception as ui_error:
                    print(f"[REG DEBUG] UI error update failed: {ui_error}")
            
            self.root.after(0, update_error_status)
            
            # Re-enable form on UI thread with safe handling
            def re_enable_form_error():
                try:
                    if hasattr(self, 'reg_employee_id') and self.reg_employee_id.winfo_exists():
                        self.reg_employee_id.configure(state='normal')
                    if hasattr(self, 'reg_employee_name') and self.reg_employee_name.winfo_exists():
                        self.reg_employee_name.configure(state='normal')
                    if hasattr(self, 'reg_department') and self.reg_department.winfo_exists():
                        self.reg_department.configure(state='normal')
                    if hasattr(self, 'capture_btn') and self.capture_btn.winfo_exists():
                        self.capture_btn.configure(state='normal', text="📷 START FACE CAPTURE")
                except Exception as form_error:
                    print(f"[REG DEBUG] Form re-enable error: {form_error}")
            
            # Resume main camera processing even on error
            def resume_main_camera_error():
                try:
                    self.main_camera_paused = False
                    print("[REG DEBUG] Main camera flag cleared after capture error")
                    
                    # Restart main camera if not active
                    if not self.camera_active:
                        print("[REG DEBUG] Restarting main camera after registration error")
                        self.start_camera()
                    
                    # Resume background face processing even on error
                    self.face_recognition.start_background_processing()
                    print("[REG DEBUG] Background face processing resumed after error")
                except Exception as camera_error:
                    print(f"[REG DEBUG] Error resuming main camera: {camera_error}")
            
            self.root.after(0, update_error_status)
            self.root.after(0, re_enable_form_error)
            self.root.after(0, resume_main_camera_error)
    
    def _capture_single_pose_improved(self, instruction: str):
        """
        Capture face vector for a single pose with quality checks
        
        Args:
            instruction: Current instruction text
            
        Returns:
            dict: Capture data with vector and metadata, or None if failed
        """
        max_attempts = 15  # Try for 3 seconds max (15 x 0.2s)
        attempts = 0
        
        while attempts < max_attempts and self.reg_capture_active and self.reg_camera_active:
            try:
                # Get frame from camera
                if hasattr(self, 'camera_manager') and self.camera_manager and self.camera_manager.cap is not None:
                    ret, frame = self.camera_manager.cap.read()
                    
                    if ret and frame is not None:
                        # Extract face vector using hybrid approach
                        face_vector, face_coords = self.face_recognition.extract_face_embedding_hybrid(frame)
                        
                        if face_vector is not None and face_coords is not None:
                            # Quality check: ensure face meets quality standards
                            if self._is_good_quality_capture_improved(frame, face_coords):
                                capture_data = {
                                    'vector': face_vector,
                                    'instruction': instruction,
                                    'timestamp': time.time(),
                                    'coords': face_coords
                                }
                                print(f"[IMPROVED REG] ✅ Captured quality vector for: {instruction}")
                                return capture_data
            
                attempts += 1
                time.sleep(0.2)  # Wait before next attempt
                
            except Exception as e:
                print(f"[IMPROVED REG] Capture error for '{instruction}': {e}")
                attempts += 1
                time.sleep(0.2)
        
        print(f"[IMPROVED REG] ❌ Failed to capture for: {instruction}")
        return None
    
    def _is_good_quality_capture_improved(self, frame, face_coords):
        """
        Check if the captured face meets quality standards
        
        Args:
            frame: Camera frame
            face_coords: Face bounding box coordinates (x, y, w, h)
            
        Returns:
            bool: True if capture meets quality standards
        """
        if face_coords is None:
            return False
        
        try:
            x, y, w, h = face_coords
            
            # Check face size (should be reasonable portion of frame)
            frame_area = frame.shape[0] * frame.shape[1]
            face_area = w * h
            face_ratio = face_area / frame_area
            
            # Face should be at least 3% of frame but not more than 40%
            if face_ratio < 0.03 or face_ratio > 0.4:
                print(f"[QUALITY CHECK] Face size ratio {face_ratio:.3f} out of range (0.03-0.4)")
                return False
            
            # Check if face is reasonably centered (not too close to edges)
            frame_h, frame_w = frame.shape[:2]
            center_x, center_y = x + w//2, y + h//2
            
            # Face center should be in middle 70% of frame
            margin_x, margin_y = frame_w * 0.15, frame_h * 0.15
            if (center_x < margin_x or center_x > frame_w - margin_x or
                center_y < margin_y or center_y > frame_h - margin_y):
                print(f"[QUALITY CHECK] Face not well centered: ({center_x}, {center_y})")
                return False
            
            # Check minimum face dimensions
            if w < 50 or h < 50:
                print(f"[QUALITY CHECK] Face too small: {w}x{h}")
                return False
            
            print(f"[QUALITY CHECK] ✅ Good quality face: size={w}x{h}, ratio={face_ratio:.3f}, center=({center_x},{center_y})")
            return True
            
        except Exception as e:
            print(f"[QUALITY CHECK] Error checking quality: {e}")
            return False

    def _check_face_distance_and_provide_feedback(self, frame, face_coords):
        """
        Check face distance and provide user feedback
        
        Args:
            frame: Camera frame
            face_coords: Face bounding box coordinates (x, y, w, h)
            
        Returns:
            tuple: (is_good_distance, feedback_message)
        """
        if face_coords is None:
            return False, "Please position your face in the camera view"
        
        try:
            x, y, w, h = face_coords
            
            # Calculate face size relative to frame
            frame_area = frame.shape[0] * frame.shape[1]
            face_area = w * h
            face_ratio = face_area / frame_area
            
            # Define distance thresholds
            TOO_FAR_THRESHOLD = 0.03    # Less than 3% of frame
            TOO_NEAR_THRESHOLD = 0.4    # More than 40% of frame
            OPTIMAL_MIN = 0.08          # 8% for optimal range
            OPTIMAL_MAX = 0.25          # 25% for optimal range
            
            # Check distance and provide specific feedback
            if face_ratio < TOO_FAR_THRESHOLD:
                return False, "Please move CLOSER to the camera"
            elif face_ratio > TOO_NEAR_THRESHOLD:
                return False, "Please move FURTHER from the camera"
            elif face_ratio < OPTIMAL_MIN:
                return True, "Move a bit closer for optimal recognition"
            elif face_ratio > OPTIMAL_MAX:
                return True, "Move a bit further for optimal recognition"
            else:
                return True, "Perfect distance - ready for recognition"
                
        except Exception as e:
            print(f"[DISTANCE CHECK] Error checking distance: {e}")
            return False, "Error checking face distance"

    def _display_distance_feedback(self, display_frame, feedback_message, is_good_distance):
        """
        Display distance feedback on the camera frame
        
        Args:
            display_frame: Frame to draw on
            feedback_message: Message to display
            is_good_distance: Whether distance is acceptable
        """
        try:
            # Choose color based on distance quality
            if is_good_distance:
                if "Perfect distance" in feedback_message:
                    color = (0, 255, 0)  # Green for perfect
                else:
                    color = (0, 255, 255)  # Yellow for acceptable
            else:
                color = (0, 0, 255)  # Red for bad distance
            
            # Display feedback at top of frame
            frame_height, frame_width = display_frame.shape[:2]
            
            # Background rectangle for better text visibility
            text_size = cv2.getTextSize(feedback_message, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)[0]
            rect_width = text_size[0] + 20
            rect_height = text_size[1] + 20
            rect_x = (frame_width - rect_width) // 2
            rect_y = 20
            
            # Draw background rectangle
            cv2.rectangle(display_frame, 
                         (rect_x, rect_y), 
                         (rect_x + rect_width, rect_y + rect_height), 
                         (0, 0, 0), -1)
            
            # Draw text
            text_x = rect_x + 10
            text_y = rect_y + text_size[1] + 10
            cv2.putText(display_frame, feedback_message, 
                       (text_x, text_y), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)
            
        except Exception as e:
            print(f"[DISPLAY FEEDBACK] Error displaying feedback: {e}")
    
    def set_attendance_mode(self, mode):
        """Set the attendance mode and switch interfaces"""
        self.attendance_mode = mode
        
        # Hide all interfaces first
        self.clock_interface.pack_forget()
        self.check_interface.pack_forget()
        
        # Reset all button colors first
        self.clock_btn.configure(fg_color=("gray70", "gray30"))
        self.check_btn.configure(fg_color=("gray70", "gray30"))
        
        # Update button colors and show appropriate interface
        if mode == "CLOCK":
            self.clock_btn.configure(fg_color=("green", "darkgreen"))
            # Show CLOCK interface
            self.clock_interface.pack(fill="both", expand=True, padx=10, pady=10)
            # Update history for CLOCK mode
            self.update_attendance_history()
        else:  # CHECK
            self.check_btn.configure(fg_color=("blue", "darkblue"))
            # Show CHECK interface
            self.check_interface.pack(fill="both", expand=True, padx=10, pady=10)
            # Update history for CHECK mode
            self.update_attendance_history()
        
        print(f"[MODE DEBUG] Switched to {mode} mode with dedicated interface")
    
    def show_manual_entry(self):
        """Show manual entry dialog based on mode"""
        if self.attendance_mode == "CHECK":
            # For CHECK mode, show selection dialog first
            self.show_check_mode_selection()
        else:
            # For CLOCK mode, use simple single-employee dialog
            self.show_simple_manual_entry()
    
    def show_check_mode_selection(self):
        """Show selection dialog for CHECK mode - Personal or Group"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Check Mode Selection")
        dialog.geometry("500x350")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (350 // 2)
        dialog.geometry(f"500x350+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='white')
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='white')
        main_frame.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="✅ CHECK MODE SELECTION",
            font=("Arial", 18, "bold"),
            bg='white',
            fg='black'
        )
        title_label.pack(pady=20)
             
        # Button frame
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=30)
        
        # Personal Check button
        personal_btn = tk.Button(
            button_frame,
            text="PERSONAL\n CHECK\n IN/OUT\n (1)",
            font=("Arial", 11, "bold"),
            bg='lightblue',
            fg='black',
            width=15,
            height=6,
            relief='raised',
            bd=2,
            command=lambda: self.handle_personal_check(dialog)
        )
        personal_btn.pack(side="left", padx=15)
        
        # Group Check button
        group_btn = tk.Button(
            button_frame,
            text="GROUP\n CHECK\n OUT\n (2)",
            font=("Arial", 11, "bold"),
            bg='lightgreen',
            fg='black',
            width=15,
            height=6,
            relief='raised',
            bd=2,
            command=lambda: self.handle_group_check(dialog)
        )
        group_btn.pack(side="right", padx=15)
        
        # Cancel button
        cancel_btn = tk.Button(
            main_frame,
            text="Cancel",
            font=("Arial", 10),
            command=dialog.destroy,
            bg='lightgray',
            fg='black',
            width=15,
            relief='raised',
            bd=1
        )
        cancel_btn.pack(pady=20)
        
        # Bind keyboard shortcuts
        dialog.bind('<Key-1>', lambda e: self.handle_personal_check(dialog))
        dialog.bind('<KP_1>', lambda e: self.handle_personal_check(dialog))
        dialog.bind('<Key-2>', lambda e: self.handle_group_check(dialog))
        dialog.bind('<KP_2>', lambda e: self.handle_group_check(dialog))
        dialog.bind('<Return>', lambda e: self.handle_personal_check(dialog))  # Enter = Personal Check
        dialog.bind('<KP_Enter>', lambda e: self.handle_personal_check(dialog))  # Numpad Enter
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Focus on dialog to capture keys
        dialog.focus_set()
    
    def handle_personal_check(self, selection_dialog):
        """Handle personal check selection"""
        selection_dialog.destroy()
        self.show_personal_check_dialog()
    
    def handle_group_check(self, selection_dialog):
        """Handle group check selection"""
        selection_dialog.destroy()
        self.show_group_checkout_dialog()
    
    def show_personal_check_dialog(self):
        """Show personal check dialog for single employee"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Personal Check In/Out")
        dialog.geometry("450x300")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"450x300+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='white')
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='white')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="👤 PERSONAL CHECK IN/OUT",
            font=("Arial", 16, "bold"),
            bg='white',
            fg='black'
        )
        title_label.pack(pady=15)
        
        # Employee ID input
        input_frame = tk.Frame(main_frame, bg='white')
        input_frame.pack(pady=15)
        
        tk.Label(
            input_frame,
            text="Employee ID:",
            font=("Arial", 12),
            bg='white',
            fg='black'
        ).pack(side="left", padx=5)
        
        # Validation for employee ID
        def validate_emp_id(char, current_text):
            return char.isdigit() and len(current_text) < 10
        
        vcmd_emp = (dialog.register(validate_emp_id), '%S', '%P')
        
        self.personal_id_entry = tk.Entry(
            input_frame,
            font=("Arial", 14),
            width=15,
            validate='key',
            validatecommand=vcmd_emp,
            justify="center"
        )
        self.personal_id_entry.pack(side="left", padx=10)
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=20)
        
        proceed_btn = tk.Button(
            button_frame,
            text="Process Check In/Out",
            font=("Arial", 12, "bold"),
            command=lambda: self.process_personal_check(dialog),
            bg='lightblue',
            fg='black',
            width=20,
            relief='raised',
            bd=2
        )
        proceed_btn.pack(side="left", padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="Cancel",
            font=("Arial", 12),
            command=dialog.destroy,
            bg='lightgray',
            fg='black',
            width=10,
            relief='raised',
            bd=1
        )
        cancel_btn.pack(side="right", padx=10)
        
        # Bind Enter key
        self.personal_id_entry.bind('<Return>', lambda e: self.process_personal_check(dialog))
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Focus on entry
        self.personal_id_entry.focus_set()
        
        # Store dialog reference for camera processing
        self.personal_dialog = dialog
        self.personal_scanning_active = True
    
    def process_personal_check(self, dialog):
        """Process personal check for single employee"""
        employee_id = self.personal_id_entry.get().strip()
        
        if not employee_id:
            self.show_temp_message("Please enter Employee ID!", "red", parent_dialog=dialog)
            return
        
        # Check if employee exists
        employee = self.db.get_employee(employee_id)
        if not employee:
            self.show_temp_message("Employee not found!", "red", parent_dialog=dialog)
            return
        
        # Close dialog and disable scanning
        dialog.destroy()
        self.personal_scanning_active = False
        
        # Process attendance directly (handles location selection if needed)
        success, message = self.process_attendance_with_location_check(employee_id, "manual")
        if success and message != "Location selection initiated":
            self.show_success_message(f"✓ {employee['name']} - {message}")
        elif not success:
            self.show_error_message(f"✗ {message}")
        # If message is "Location selection initiated", the location dialog will handle the UI
    
    def show_simple_manual_entry(self):
        """Show simple manual entry dialog for CLOCK mode"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Manual Entry")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"400x250+{x}+{y}")
        
        # Entry widgets
        tk.Label(dialog, text="Enter Employee ID:", font=("Arial", 16)).pack(pady=20)
        
        # Create validation function for numbers only with length limit
        def validate_input(char, current_text):
            """Validate that input contains only numbers and is within length limit"""
            # Allow only digits and limit length to 10 characters
            return char.isdigit() and len(current_text) < 10
        
        # Register the validation function
        vcmd = (dialog.register(validate_input), '%S', '%P')
        
        entry = tk.Entry(dialog, font=("Arial", 14), width=20, justify="center",
                        validate='key', validatecommand=vcmd)
        entry.pack(pady=10)
        entry.focus_set()
        
        def submit():
            employee_id = entry.get().strip()
            dialog.destroy()
            if employee_id and employee_id.isdigit():
                self.process_manual_entry(employee_id)
            elif employee_id:
                self.show_error_message("✗ Employee ID must contain only numbers")
            else:
                self.show_error_message("✗ Please enter an Employee ID")
        
        def cancel():
            dialog.destroy()
        
        # Buttons
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Submit", font=("Arial", 12), 
                 command=submit, width=10).pack(side="left", padx=10)
        tk.Button(button_frame, text="Cancel", font=("Arial", 12), 
                 command=cancel, width=10).pack(side="left", padx=10)
        
        # Bind Enter key
        entry.bind('<Return>', lambda e: submit())
        dialog.bind('<Escape>', lambda e: cancel())
    
    def process_manual_entry(self, employee_id):
        """Process manual employee ID entry"""
        employee = self.db.get_employee(employee_id)
        if employee:
            success, message = self.process_attendance_with_location_check(employee_id, "manual")
            if success and message != "Location selection initiated":
                self.show_success_message(f"✓ {employee['name']} - {message}")
            elif not success:
                # Check if this is an early clock-out error
                if self.is_early_clockout_error(message):
                    self.handle_early_clockout_error(employee['name'], message)
                else:
                    self.show_error_message(f"✗ {message}")
            # If message is "Location selection initiated", the location dialog will handle the UI
        else:
            self.show_employee_not_found_dialog(employee_id)
    
    def process_attendance_with_location_check(self, employee_id, method):
        """Process attendance with proper location handling for CHECK OUT"""
        employee = self.db.get_employee(employee_id)
        if not employee:
            return False, f"Employee {employee_id} not found"
        
        # For CHECK mode, check if this will be a checkout (need location first)
        if self.attendance_mode == "CHECK":
            # Check what the next action would be
            today_records = self.attendance_manager.get_employee_attendance_today(employee_id)
            
            # Filter for check records and sort by timestamp to find the most recent
            check_records = [r for r in today_records if r.get('attendance_type') == 'check']
            check_records.sort(key=lambda x: x['timestamp'], reverse=True)  # Most recent first
            
            last_check_record = check_records[0] if check_records else None
            
            # Determine if this will be a checkout
            will_be_checkout = False
            if not last_check_record:
                # No check records today - default is checkout
                will_be_checkout = True
            else:
                # Will checkout if last status was 'in'
                will_be_checkout = (last_check_record['status'] == 'in')
            
            if will_be_checkout:
                # Show location selection first, then process attendance
                self.handle_checkout_with_location_first(employee_id, method)
                return True, "Location selection initiated"
        
        # Normal processing for CLOCK mode or CHECK IN
        success, message = self.attendance_manager.process_attendance(
            employee_id, method, self.attendance_mode, None
        )
        return success, message
    
    def handle_checkout_with_location_first(self, employee_id, method):
        """Handle checkout by showing location selection first, then recording attendance"""
        def on_location_selected(location):
            if location:
                # Location selected - proceed with checkout and record location
                current_time = self.attendance_manager.get_current_time()
                record_id = self.db.record_attendance(employee_id, method, "out", "check", current_time)
                
                if record_id:
                    # Prepare location data for attendance record
                    location_data = {
                        "location_name": location.get('name', ''),
                        "address": location.get('address', '')
                    }
                    
                    # Update the attendance record with location information
                    success = self.db.update_attendance_location(record_id, location_data)
                    
                    if success:
                        # Also save to location manager for favorites/history (optional)
                        self.location_manager.save_checkout_location(
                            record_id, employee_id, location
                        )
                        
                        employee = self.db.get_employee(employee_id)
                        location_name = location.get('name', 'Selected location')
                        self.show_success_message(
                            f"✓ {employee['name']} checked out\n📍 Going to: {location_name}"
                        )
                        # Update attendance history
                        self.update_attendance_history()
                    else:
                        self.show_error_message("✗ Failed to save location")
                else:
                    self.show_error_message("✗ Failed to record checkout")
            else:
                # User cancelled location selection - do not record checkout
                employee = self.db.get_employee(employee_id)
                self.show_error_message(f"✗ {employee['name']} checkout cancelled - location required")
        
        # Open location selector dialog
        from core.location_selector import LocationSelector
        LocationSelector(
            parent=self.root,
            employee_id=employee_id,
            callback=on_location_selected
        )
    
    def handle_checkout_location_selection(self, attendance_id, employee_id):
        """Handle location selection for check-out (legacy method for existing attendance records)"""
        def on_location_selected(location):
            if location:
                # Prepare location data for attendance record
                location_data = {
                    "location_name": location.get('name', ''),
                    "address": location.get('address', '')
                }
                
                # Update the attendance record with location information
                success = self.db.update_attendance_location(attendance_id, location_data)
                
                if success:
                    # Also save to location manager for favorites/history (optional)
                    self.location_manager.save_checkout_location(
                        attendance_id, employee_id, location
                    )
                    
                    employee = self.db.get_employee(employee_id)
                    location_name = location.get('name', 'Selected location')
                    self.show_success_message(
                        f"✓ {employee['name']} checked out\n📍 Going to: {location_name}"
                    )
                else:
                    self.show_error_message("✗ Failed to save location")
            else:
                # User cancelled location selection - that's okay, checkout is still recorded
                employee = self.db.get_employee(employee_id)
                self.show_success_message(f"✓ {employee['name']} checked out")
        
        # Open location selector dialog
        from core.location_selector import LocationSelector
        LocationSelector(
            parent=self.root,
            employee_id=employee_id,
            callback=on_location_selected
        )
    
    def show_employee_not_found_dialog(self, employee_id):
        """Show a dedicated error dialog for employee not found"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Employee Not Found")
        dialog.geometry("350x200")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (dialog.winfo_screenheight() // 2) - (200 // 2)
        dialog.geometry(f"350x200+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='#ff4444')
        
        # Error icon and message
        main_frame = tk.Frame(dialog, bg='#ff4444')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Error icon
        tk.Label(main_frame, text="❌", font=("Arial", 32), 
                bg='#ff4444', fg='white').pack(pady=10)
        
        # Error message
        tk.Label(main_frame, text="EMPLOYEE NOT FOUND", 
                font=("Arial", 14, "bold"), bg='#ff4444', fg='white').pack()
        
        tk.Label(main_frame, text=f"Employee ID: {employee_id}", 
                font=("Arial", 12), bg='#ff4444', fg='white').pack(pady=5)
        
        tk.Label(main_frame, text="Please check the ID and try again", 
                font=("Arial", 10), bg='#ff4444', fg='white').pack(pady=5)
        
        # OK button
        ok_button = tk.Button(main_frame, text="OK", font=("Arial", 12, "bold"),
                             command=dialog.destroy, width=10,
                             bg='white', fg='#ff4444', relief='raised', bd=2)
        ok_button.pack(pady=15)
        ok_button.focus_set()
        
        # Bind Enter and Escape keys
        dialog.bind('<Return>', lambda e: dialog.destroy())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Auto-close after 5 seconds
        dialog.after(5000, dialog.destroy)
    
    def show_already_checked_out_dialog(self, employee_name, employee_id):
        """Show error dialog for employee already checked out"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Already Checked Out")
        dialog.geometry("350x200")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (dialog.winfo_screenheight() // 2) - (200 // 2)
        dialog.geometry(f"350x200+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='#ff8800')
        
        # Warning icon and message
        main_frame = tk.Frame(dialog, bg='#ff8800')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Warning icon
        tk.Label(main_frame, text="⚠️", font=("Arial", 32), 
                bg='#ff8800', fg='white').pack(pady=10)
        
        # Warning message
        tk.Label(main_frame, text="ALREADY CHECKED OUT", 
                font=("Arial", 14, "bold"), bg='#ff8800', fg='white').pack()
        
        tk.Label(main_frame, text=f"{employee_name} (ID: {employee_id})", 
                font=("Arial", 12), bg='#ff8800', fg='white').pack(pady=5)
        
        tk.Label(main_frame, text="Employee is already in checked out status", 
                font=("Arial", 10), bg='#ff8800', fg='white').pack(pady=5)
        
        # OK button
        ok_button = tk.Button(main_frame, text="OK", font=("Arial", 12, "bold"),
                             command=dialog.destroy, width=10,
                             bg='white', fg='#ff8800', relief='raised', bd=2)
        ok_button.pack(pady=15)
        ok_button.focus_set()
        
        # Bind Enter and Escape keys
        dialog.bind('<Return>', lambda e: dialog.destroy())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Auto-close after 5 seconds
        dialog.after(5000, dialog.destroy)
    
    def show_group_checkout_dialog(self):
        """Show the group checkout dialog for collecting employee IDs"""
        dialog = tk.Toplevel(self.root)
        dialog.title("CHECK Mode - Manual Entry")
        dialog.geometry("600x700")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"600x700+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='white')
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='white')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="GROUP CHECK OUT",
            font=("Arial", 20, "bold"),
            bg='white',
            fg='black'
        )
        title_label.pack(pady=10)
        
        # Employee IDs collection
        ids_frame = tk.Frame(main_frame, bg='white')
        ids_frame.pack(fill="both", expand=True, pady=10)
        
        # Employee IDs list frame
        self.group_ids_frame = tk.Frame(ids_frame, bg='white', relief='sunken', bd=2)
        self.group_ids_frame.pack(fill="both", expand=True, pady=10, padx=20)
        
        # Scrollable area for employee list
        canvas = tk.Canvas(self.group_ids_frame, bg='white', height=200)
        scrollbar = tk.Scrollbar(self.group_ids_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg='white')
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Current employee input
        input_frame = tk.Frame(ids_frame, bg='white')
        input_frame.pack(fill="x", pady=5)
        
        tk.Label(
            input_frame,
            text="Employee ID:",
            font=("Arial", 10),
            bg='white',
            fg='gray'
        ).pack(side="left", padx=5)
        
        # Validation for employee ID
        def validate_emp_id(char, current_text):
            return char.isdigit() and len(current_text) < 10
        
        vcmd_emp = (dialog.register(validate_emp_id), '%S', '%P')
        
        self.group_current_id = tk.Entry(
            input_frame,
            font=("Arial", 12),
            width=15,
            validate='key',
            validatecommand=vcmd_emp
        )
        self.group_current_id.pack(side="left", padx=5)
        
        add_btn = tk.Button(
            input_frame,
            text="Add",
            font=("Arial", 10),
            command=lambda: self.add_employee_to_group(),
            bg='#f0f0f0',
            fg='black',
            relief='raised',
            bd=2
        )
        add_btn.pack(side="left", padx=5)
        
        # Initialize group data
        self.group_employees = []
        self.group_employee_labels = []
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(fill="x", pady=20)
        
        proceed_btn = tk.Button(
            button_frame,
            text="Proceed to Location Selection",
            font=("Arial", 12, "bold"),
            command=lambda: self.proceed_group_location_selection(dialog),
            bg='#f0f0f0',
            fg='black',
            relief='raised',
            bd=2,
            width=25
        )
        proceed_btn.pack(side="left", padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="Cancel",
            font=("Arial", 12),
            command=dialog.destroy,
            bg='#f0f0f0',
            fg='black',
            relief='raised',
            bd=2,
            width=10
        )
        cancel_btn.pack(side="right", padx=10)
        
        # Bind Enter key for adding employees
        self.group_current_id.bind('<Return>', lambda e: self.add_employee_to_group())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Focus on employee ID entry initially
        self.group_current_id.focus_set()
        
        # Store dialog reference for camera processing
        self.group_dialog = dialog
        
        # Enable camera scanning for this dialog
        self.group_scanning_active = True
    
    def add_employee_to_group(self):
        """Add an employee to the group checkout list"""
        employee_id = self.group_current_id.get().strip()
        
        if not employee_id:
            return
        
        # Check if employee exists
        employee = self.db.get_employee(employee_id)
        if not employee:
            # Use same error dialog as CLOCK mode
            self.show_employee_not_found_dialog(employee_id)
            return
        
        # Check if employee is already checked out
        today_records = self.attendance_manager.get_employee_attendance_today(employee_id)
        check_records = [r for r in today_records if r.get('attendance_type') == 'check']
        check_records.sort(key=lambda x: x['timestamp'], reverse=True)  # Most recent first
        
        last_check_record = check_records[0] if check_records else None
        
        # If last check record is 'out', employee is already checked out
        if last_check_record and last_check_record['status'] == 'out':
            self.show_already_checked_out_dialog(employee['name'], employee_id)
            return
        
        # Check if already added to current group
        if employee_id in [emp['employee_id'] for emp in self.group_employees]:
            self.show_temp_message("Employee already added to group!", "orange")
            return
        
        # Add employee
        self.group_employees.append({
            'employee_id': employee_id,
            'name': employee['name']
        })
        
        # Update display
        self.update_group_employee_display()
        
        # Clear input
        self.group_current_id.delete(0, tk.END)
        
        print(f"[GROUP DEBUG] Added employee: {employee['name']} ({employee_id})")
    
    def update_group_employee_display(self):
        """Update the display of employees in the group"""
        # Clear existing labels
        for label in self.group_employee_labels:
            label.destroy()
        self.group_employee_labels.clear()
        
        # Add new labels
        for i, emp in enumerate(self.group_employees):
            emp_frame = tk.Frame(self.scrollable_frame, bg='white')
            emp_frame.pack(fill="x", padx=5, pady=2)
            
            emp_label = tk.Label(
                emp_frame,
                text=f"{i+1}. {emp['name']} (ID: {emp['employee_id']})",
                font=("Arial", 10),
                bg='white',
                anchor='w'
            )
            emp_label.pack(side="left", fill="x", expand=True)
            
            remove_btn = tk.Button(
                emp_frame,
                text="✕",
                font=("Arial", 8),
                command=lambda idx=i: self.remove_employee_from_group(idx),
                bg='red',
                fg='white',
                width=3
            )
            remove_btn.pack(side="right")
            
            self.group_employee_labels.extend([emp_frame, emp_label, remove_btn])
    
    def remove_employee_from_group(self, index):
        """Remove an employee from the group"""
        if 0 <= index < len(self.group_employees):
            removed_emp = self.group_employees.pop(index)
            self.update_group_employee_display()
            print(f"[GROUP DEBUG] Removed employee: {removed_emp['name']} ({removed_emp['employee_id']})")
    
    def show_temp_message(self, message, color, parent_dialog=None):

        if parent_dialog is not None:
            try:
                parent_dialog.destroy()
            except:
                pass
    
        """Show a temporary message dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Information")
        dialog.geometry("350x180")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (dialog.winfo_screenheight() // 2) - (180 // 2)
        dialog.geometry(f"350x180+{x}+{y}")
        
        # Configure dialog background based on color
        bg_color = '#ff8800' if color == 'orange' else '#ff4444'
        dialog.configure(bg=bg_color)
        
        # Message frame
        main_frame = tk.Frame(dialog, bg=bg_color)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Icon based on color
        icon = "⚠️" if color == 'orange' else "❌"
        tk.Label(main_frame, text=icon, font=("Arial", 32), 
                bg=bg_color, fg='white').pack(pady=10)
        
        # Message
        tk.Label(main_frame, text=message, 
                font=("Arial", 12, "bold"), bg=bg_color, fg='white').pack(pady=5)
        
        # OK button
        ok_button = tk.Button(main_frame, text="OK", font=("Arial", 12, "bold"),
                             command=dialog.destroy, width=10,
                             bg='white', fg=bg_color, relief='raised', bd=2)
        ok_button.pack(pady=15)
        ok_button.focus_set()
        
        # Bind Enter and Escape keys
        dialog.bind('<Return>', lambda e: dialog.destroy())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

        dialog.focus_set()
        
        # Auto-close after 3 seconds
        dialog.after(3000, dialog.destroy)
    
    def proceed_group_location_selection(self, dialog):
        """Proceed to location selection for the group"""
        if not self.group_employees:
            self.show_temp_message("Add at least one employee!", "red")
            return
        
        # Close the group dialog
        dialog.destroy()
        self.group_scanning_active = False
        
        # Show location selector for the group
        self.show_group_location_selector()
    
    def show_group_location_selector(self):
        """Show location selector for group checkout"""
        def on_location_selected(location):
            if location:
                self.process_group_checkout(location)
            else:
                self.show_error_message("✗ Group checkout cancelled - location required")
        
        # Open location selector dialog
        from core.location_selector import LocationSelector
        LocationSelector(
            parent=self.root,
            employee_id=f"GROUP_{len(self.group_employees)}",  # Group identifier
            callback=on_location_selected
        )
    
    def process_group_checkout(self, location):
        """Process checkout for all employees in the group"""
        successful_checkouts = []
        failed_checkouts = []
        
        current_time = self.attendance_manager.get_current_time()
        
        for emp in self.group_employees:
            employee_id = emp['employee_id']
            
            # Record attendance
            record_id = self.db.record_attendance(employee_id, "manual", "out", "check", current_time)
            
            if record_id:
                # Prepare location data for attendance record
                location_data = {
                    "location_name": location.get('name', ''),
                    "address": location.get('address', '')
                }
                
                # Update the attendance record with location information
                success = self.db.update_attendance_location(record_id, location_data)
                
                if success:
                    successful_checkouts.append(emp['name'])
                    
                    # Also save to location manager for favorites/history
                    self.location_manager.save_checkout_location(
                        record_id, employee_id, location
                    )
                else:
                    failed_checkouts.append(emp['name'])
            else:
                failed_checkouts.append(emp['name'])
        
        # Show results
        location_name = location.get('name', 'Selected location')
        if successful_checkouts:
            if len(successful_checkouts) == 1:
                # Single employee checkout
                success_msg = f"✓ {successful_checkouts[0]} checked out\n📍 Going to: {location_name}"
            else:
                # Multiple employee checkout
                success_names = ', '.join(successful_checkouts)
                success_msg = f"✓ Group checkout successful!\n👥 {len(successful_checkouts)} employees\n📍 Going to: {location_name}\n\nEmployees: {success_names}"
            
            if failed_checkouts:
                fail_names = ', '.join(failed_checkouts)
                success_msg += f"\n\n⚠️ Failed: {fail_names}"
            self.show_success_message(success_msg)
        else:
            self.show_error_message("✗ Checkout failed for all employees")
        
        # Clear group data
        self.group_employees.clear()
        
        # Update attendance history
        self.update_attendance_history()
        
        print(f"[GROUP DEBUG] Group checkout completed - Success: {len(successful_checkouts)}, Failed: {len(failed_checkouts)}")
    
    def show_qr_recognition_confirmation(self, qr_code, employee):
        """Show confirmation dialog for QR code recognition"""
        try:
            print(f"[CONFIRMATION DEBUG] Starting QR confirmation dialog for {employee['name']}")
            
            # Turn OFF camera after recognition (manual mode)
            self.stop_camera()
            print(f"[CONFIRMATION DEBUG] Camera turned OFF after QR recognition")
            
            # Create confirmation dialog using messagebox for reliability
            employee_role = employee.get('role', 'Staff')
            role_icon = '🛡️' if employee_role == 'Security' else '👥'
            message = f"QR Code Recognition Confirmation\n\nScanned: {employee['name']}\nEmployee ID: {employee['employee_id']}\nRole: {role_icon} {employee_role}\nQR Data: {qr_code.get('data', 'N/A')}\n\nIs this scan correct?"
            
            print(f"[CONFIRMATION DEBUG] Showing messagebox")
            result = messagebox.askyesno("Confirm QR Code Recognition", message)
            print(f"[CONFIRMATION DEBUG] User response: {result}")
            
            # Process result
            if result:
                # User confirmed - process attendance
                success, message = self.process_attendance_with_location_check(
                    employee['employee_id'], "qr_code"
                )
                
                if success:
                    print(f"[QR SCAN DEBUG] Attendance success: {message}")
                    if message != "Location selection initiated":
                        self.show_success_message(f"✓ {employee['name']} - {message}")
                else:
                    print(f"[QR SCAN DEBUG] Attendance failed: {message}")
                    # Check if this is an early clock-out error
                    if self.is_early_clockout_error(message):
                        self.handle_early_clockout_error(employee['name'], message)
                    else:
                        self.show_error_message(f"✗ {message}")
                
                # Keep camera OFF after successful processing
                print(f"[CONFIRMATION DEBUG] Keeping camera OFF after successful QR processing")
                
            else:
                # User rejected - keep camera OFF, let them manually try again
                print(f"[QR SCAN DEBUG] QR scan rejected by user: {employee['name']}")
                self.show_error_message("❌ QR code scan rejected\n\nPress + to try scanning again")
                
                # Keep camera OFF - employee must manually press + to try again
                print(f"[CONFIRMATION DEBUG] Camera remains OFF - manual activation required for retry")
                
        except Exception as e:
            print(f"Error in QR recognition confirmation: {e}")
            print(f"[CONFIRMATION DEBUG] Exception traceback: {traceback.format_exc()}")
            # Keep camera OFF in case of error - manual activation required
            self.show_error_message("Error in QR code confirmation\n\nPress + to try again")

    def show_face_recognition_confirmation(self, face, employee):
        """Show confirmation dialog for face recognition"""
        try:
            print(f"[CONFIRMATION DEBUG] Starting face confirmation dialog for {employee['name']}")
            
            # Turn OFF camera after recognition (manual mode)
            self.stop_camera()
            print(f"[CONFIRMATION DEBUG] Camera turned OFF after face recognition")
            
            # Create confirmation dialog using messagebox for reliability
            employee_role = employee.get('role', 'Staff')
            role_icon = '🛡️' if employee_role == 'Security' else '👥'
            message = f"Face Recognition Confirmation\n\nRecognized: {face['name']}\nEmployee ID: {face['employee_id']}\nRole: {role_icon} {employee_role}\nConfidence: {face.get('confidence', 0.0)*100:.2f}%\n\nIs this recognition correct?"

            print(f"[CONFIRMATION DEBUG] Showing messagebox")
            result = messagebox.askyesno("Confirm Face Recognition", message)
            print(f"[CONFIRMATION DEBUG] User response: {result}")
            
            # Process result
            if result:
                # User confirmed - process attendance
                success, message = self.process_attendance_with_location_check(
                    face['employee_id'], "face_recognition"
                )
                
                if success:
                    print(f"[FACE DEBUG] Attendance success: {message}")
                    if message != "Location selection initiated":
                        self.show_success_message(f"✓ {employee['name']} - {message}")
                else:
                    print(f"[FACE DEBUG] Attendance failed: {message}")
                    # Check if this is an early clock-out error
                    if self.is_early_clockout_error(message):
                        self.handle_early_clockout_error(employee['name'], message)
                    else:
                        self.show_error_message(f"✗ {message}")
                
                # Keep camera OFF after successful processing
                print(f"[CONFIRMATION DEBUG] Keeping camera OFF after successful face processing")
                
            else:
                # User rejected - keep camera OFF, let them manually try again
                print(f"[FACE DEBUG] Recognition rejected by user: {face['name']}")
                self.show_error_message("❌ Face recognition rejected\n\nPress + to try recognition again")
                
                # Keep camera OFF - employee must manually press + to try again
                print(f"[CONFIRMATION DEBUG] Camera remains OFF - manual activation required for retry")
                
        except Exception as e:
            print(f"Error in face recognition confirmation: {e}")
            print(f"[CONFIRMATION DEBUG] Exception traceback: {traceback.format_exc()}")
            # Keep camera OFF in case of error - manual activation required
            self.show_error_message("Error in face recognition confirmation\n\nPress + to try again")

    def show_success_message(self, message):
        """Show success message"""
        self.message_frame.pack(fill="x", padx=30, pady=10)
        self.message_label.configure(text=message, text_color="green")
        self.status_label.configure(text="● SUCCESS", text_color="green")
        
        # Update attendance history when someone scans
        self.update_attendance_history()
        
        # Auto-hide after timeout
        self.root.after(self.auto_timeout * 1000, self.clear_message)
    
    def show_error_message(self, message):
        """Show error message"""
        self.message_frame.pack(fill="x", padx=30, pady=10)
        self.message_label.configure(text=message, text_color="red")
        self.status_label.configure(text="● ERROR", text_color="red")
        
        # Auto-hide after timeout
        self.root.after(self.auto_timeout * 1000, self.clear_message)
    
    def show_early_clockout_error(self, employee_name, min_time, shift_name):
        """Show dedicated error dialog for early clock-out attempts"""
        # Create a modal dialog
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Clock-Out Restricted")
        dialog.geometry("500x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"500x300+{x}+{y}")
        
        # Configure dialog
        dialog.configure(fg_color=("gray90", "gray10"))
        dialog.resizable(False, False)
        
        # Main frame
        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Error icon and title
        title_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        title_frame.pack(fill="x", pady=(10, 20))
        
        error_label = ctk.CTkLabel(
            title_frame,
            text="⚠️ CLOCK-OUT RESTRICTED",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="red"
        )
        error_label.pack()
        
        # Employee name
        name_label = ctk.CTkLabel(
            main_frame,
            text=f"Employee: {employee_name}",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        name_label.pack(pady=(0, 15))
        
        # Main message
        message_label = ctk.CTkLabel(
            main_frame,
            text=f"You cannot clock out before {min_time}",
            font=ctk.CTkFont(size=14),
            text_color="red"
        )
        message_label.pack(pady=(0, 10))
        
        # Shift info
        shift_label = ctk.CTkLabel(
            main_frame,
            text=f"Current Shift: {shift_name}",
            font=ctk.CTkFont(size=12)
        )
        shift_label.pack(pady=(0, 20))
        
        # Current time display
        current_time = datetime.now().strftime("%I:%M %p")
        time_label = ctk.CTkLabel(
            main_frame,
            text=f"Current Time: {current_time}",
            font=ctk.CTkFont(size=12)
        )
        time_label.pack(pady=(0, 20))
        
        # OK button
        ok_button = ctk.CTkButton(
            main_frame,
            text="OK",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=100,
            height=35,
            command=dialog.destroy
        )
        ok_button.pack(pady=(10, 10))
        
        # Auto-focus the OK button
        ok_button.focus_set()
        
        # Bind Enter key to close dialog
        dialog.bind('<Return>', lambda e: dialog.destroy())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
    
    def is_early_clockout_error(self, message):
        """Check if error message is about early clock-out"""
        return "Cannot clock out before" in message
    
    def handle_early_clockout_error(self, employee_name, message):
        """Handle early clock-out error with custom dialog"""
        # Parse the error message to extract time and shift info
        # Message format: "Cannot clock out before 5:00 PM (Early Shift)"
        import re
        match = re.search(r'Cannot clock out before (\d{1,2}:\d{2} [AP]M) \(([^)]+)\)', message)
        
        if match:
            min_time = match.group(1)
            shift_name = match.group(2)
        else:
            # Fallback if parsing fails
            min_time = "5:00 PM"
            shift_name = "Shift"
        
        # Show the custom dialog
        self.show_early_clockout_error(employee_name, min_time, shift_name)
    
    def clear_message(self):
        """Clear status message"""
        self.message_frame.pack_forget()
        self.status_label.configure(
            text="Please position yourself in front of the camera",
            text_color="green"
        )
    
    def load_known_faces(self):
        """Load known faces from database (now using face vectors)"""
        print("[FACE DEBUG] Loading known faces from database...")
        
        # Try to load face vectors first (new method)
        employees_with_vectors = self.db.get_all_face_vectors()
        
        if employees_with_vectors:
            print(f"[FACE DEBUG] Found {len(employees_with_vectors)} employees with face vectors")
            for emp in employees_with_vectors:
                print(f"[FACE DEBUG] - {emp['name']} ({emp['employee_id']}): Vector loaded")
            # Pass the database manager instead of the list
            self.face_recognition.load_known_faces(self.db)
        else:
            # Fallback to old image-based method for backward compatibility
            print("[FACE DEBUG] No face vectors found, falling back to face images...")
            employees_with_faces = self.db.get_all_face_images()
            print(f"[FACE DEBUG] Found {len(employees_with_faces)} employees with face images")
            
            if employees_with_faces:
                print("[FACE DEBUG] ⚠️  Consider running migration to convert images to vectors!")
                for emp in employees_with_faces:
                    print(f"[FACE DEBUG] - {emp['name']} ({emp['employee_id']}): {emp.get('face_image_path', 'No image path')}")
                
                # For backward compatibility, we'll need to convert images to vectors on-the-fly
                # This is less efficient but ensures the system continues working
                self.convert_and_load_legacy_faces(employees_with_faces)
    
    def convert_and_load_legacy_faces(self, employees_with_faces):
        """Convert legacy face images to vectors on-the-fly (for backward compatibility)"""
        converted_employees = []
        
        for emp in employees_with_faces:
            try:
                # Load the image file and extract embedding
                import cv2
                img = cv2.imread(emp['face_image_path'])
                if img is not None:
                    face_vector, _ = self.face_recognition.extract_face_embedding_hybrid(img)
                    if face_vector is not None:
                        converted_employees.append({
                            'employee_id': emp['employee_id'],
                            'name': emp['name'],
                            'face_vector': face_vector
                        })
                        print(f"[FACE DEBUG] ✓ Converted {emp['name']} image to vector")
                    else:
                        print(f"[FACE DEBUG] ✗ Failed to extract face from {emp['name']} image")
                else:
                    print(f"[FACE DEBUG] ✗ Could not load image file for {emp['name']}")
            except Exception as e:
                print(f"[FACE DEBUG] ✗ Failed to convert {emp['name']} image: {e}")
        
        if converted_employees:
            # Store converted employees in database first, then load them properly
            for emp in converted_employees:
                # Convert and save each employee's face vector
                try:
                    success = self.db.add_employee_with_face_vector(
                        employee_id=emp['employee_id'],
                        name=emp['name'],
                        face_vector=emp['face_vector'],
                        department=emp.get('department', 'Unknown')
                    )
                    if success:
                        print(f"[FACE DEBUG] ✓ Migrated {emp['name']} to vector format")
                except Exception as e:
                    print(f"[FACE DEBUG] ✗ Failed to migrate {emp['name']}: {e}")
            
            # Now load from database properly
            self.face_recognition.load_known_faces(self.db)
    
    def update_attendance_history(self):
        """Update the attendance history display for the current mode"""
        # Determine which history frame to update based on current mode
        if self.attendance_mode == "CLOCK":
            history_frame = self.clock_history_frame
            mode_filter = "clock"
        else:  # CHECK
            history_frame = self.check_history_frame
            mode_filter = "check"
        
        # Clear existing history
        for widget in history_frame.winfo_children():
            widget.destroy()
        
        # Get today's attendance with explicit date check
        attendance_records = self.db.get_attendance_today()
        
        # Debug: Show what date we're filtering for
        current_date = datetime.now().strftime("%Y-%m-%d")
        print(f"[ATTENDANCE DEBUG] Looking for {mode_filter} records on: {current_date}")
        print(f"[ATTENDANCE DEBUG] Found {len(attendance_records)} total records")
        
        # Filter for today's records and current mode
        mode_records = []
        for record in attendance_records:
            record_date = record['timestamp'][:10]  # Extract YYYY-MM-DD part
            if record_date == current_date and record.get('attendance_type') == mode_filter:
                mode_records.append(record)
        
        print(f"[ATTENDANCE DEBUG] After filtering for {mode_filter} mode: {len(mode_records)} records")
        
        if not mode_records:
            no_records_label = ctk.CTkLabel(
                history_frame,
                text=f"No {mode_filter.upper()} records for {current_date}",
                font=ctk.CTkFont(size=16),
                text_color="gray"
            )
            no_records_label.pack(pady=20)
            return
        
        # Group records by employee and keep all records (not just latest)
        employee_records = {}
        for record in mode_records:  # Use filtered records
            emp_id = record['employee_id']
            if emp_id not in employee_records:
                employee_records[emp_id] = {
                    'name': record['name'],
                    'records': []
                }
            employee_records[emp_id]['records'].append(record)
        
        # Sort records by time for each employee (newest first)
        for emp_id in employee_records:
            employee_records[emp_id]['records'].sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Display all records for each employee
        for emp_id, data in employee_records.items():
            self.create_employee_history_section(emp_id, data, history_frame)
    
    def create_employee_history_section(self, employee_id, data, history_frame):
        """Create a section showing all attendance records for one employee"""
        # Main frame for this employee - enhanced styling with more space
        border_color = ("green", "lightgreen") if self.attendance_mode == "CLOCK" else ("blue", "lightblue")
        main_frame = ctk.CTkFrame(
            history_frame, 
            border_width=2, 
            border_color=border_color,
            fg_color=("gray95", "gray15")
        )
        main_frame.pack(fill="x", padx=8, pady=12)
        
        # Employee header - larger and more prominent
        header_frame = ctk.CTkFrame(main_frame, fg_color=("blue", "darkblue"))
        header_frame.pack(fill="x", padx=6, pady=6)
        
        # Employee info in horizontal layout
        info_container = ctk.CTkFrame(header_frame, fg_color="transparent")
        info_container.pack(fill="x", padx=12, pady=8)
        
        # Employee name and ID - larger font
        name_label = ctk.CTkLabel(
            info_container,
            text=f"👤 {data['name']}",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="white",
            anchor="w"
        )
        name_label.pack(side="left", fill="x", expand=True)
        
        # Employee ID badge
        id_label = ctk.CTkLabel(
            info_container,
            text=f"ID: {employee_id}",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="lightgray",
            anchor="e"
        )
        id_label.pack(side="right", padx=10)
        
        # Current status (based on most recent record) - larger and more prominent
        latest_record = data['records'][0]  # Most recent record (first in descending order)
        
        # Determine status text based on attendance type and status
        attendance_type = latest_record.get('attendance_type', 'check')
        if attendance_type == 'clock':
            if latest_record['status'] == 'in':
                current_status = "LATE CLOCKED IN" if latest_record.get('late', False) else "CLOCKED IN"
            else:
                current_status = "CLOCKED OUT"
        else:  # check
            current_status = "CHECKED IN" if latest_record['status'] == 'in' else "CHECKED OUT"
        
        status_color = "lightgreen" if latest_record['status'] == 'in' else "red"
        status_icon = "🟢" if latest_record['status'] == 'in' else "🔴"
        
        # Status in a separate frame for better visibility
        status_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        status_frame.pack(fill="x", padx=12, pady=5)
        
        # Check if this is a late status to highlight "LATE" in red
        if latest_record.get('late', False) and attendance_type == 'clock' and latest_record['status'] == 'in':
            # Create horizontal frame for mixed color text
            mixed_status_frame = ctk.CTkFrame(status_frame, fg_color="transparent")
            mixed_status_frame.pack(pady=3)
            
            # Status prefix
            prefix_label = ctk.CTkLabel(
                mixed_status_frame,
                text=f"{status_icon} CURRENT STATUS: ",
                font=ctk.CTkFont(size=18, weight="bold"),
                text_color=status_color
            )
            prefix_label.pack(side="left")
            
            # LATE in red
            late_label = ctk.CTkLabel(
                mixed_status_frame,
                text="LATE ",
                font=ctk.CTkFont(size=18, weight="bold"),
                text_color="red"
            )
            late_label.pack(side="left")
            
            # Rest of status
            rest_label = ctk.CTkLabel(
                mixed_status_frame,
                text="CLOCKED IN",
                font=ctk.CTkFont(size=18, weight="bold"),
                text_color=status_color
            )
            rest_label.pack(side="left")
        else:
            # Normal single-color status
            status_label = ctk.CTkLabel(
                status_frame,
                text=f"{status_icon} CURRENT STATUS: {current_status}",
                font=ctk.CTkFont(size=18, weight="bold"),
                text_color=status_color,
                anchor="center"
            )
            status_label.pack(pady=3)
        
        # Records list
        records_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        records_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # Show all records for this employee (chronological order)
        for i, record in enumerate(data['records']):
            self.create_single_record_entry(records_frame, record, i)
    
    def create_single_record_entry(self, parent, record, index):
        """Create a single attendance record entry - enhanced for larger display"""
        # Record frame with better styling and more space
        bg_color = ("gray90", "gray20") if index % 2 == 0 else ("white", "gray25")
        record_frame = ctk.CTkFrame(parent, fg_color=bg_color, height=50)
        record_frame.pack(fill="x", pady=3, padx=5)
        record_frame.pack_propagate(False)
        
        # Parse timestamp
        record_time = datetime.strptime(record['timestamp'], "%Y-%m-%d %H:%M:%S")
        time_str = record_time.strftime("%H:%M:%S")
        date_str = record_time.strftime("%b %d")
        
        # Status indicator
        status_text = "IN" if record['status'] == 'in' else "OUT"
        status_color = "lightgreen" if record['status'] == 'in' else "red"
        status_icon = "🟢" if record['status'] == 'in' else "🔴"
        
        # Method indicator with icons
        method_icons = {
            'face_recognition': 'Face Recognition',
            'qr_code': 'QR/Barcode', 
            'manual': 'Manual Entry'
        }
        method_text = method_icons.get(record['method'], f"📋 {record['method']}")
        
        # Create main horizontal layout with better spacing
        info_frame = ctk.CTkFrame(record_frame, fg_color="transparent")
        info_frame.pack(fill="both", expand=True, padx=15, pady=8)
        
        # Time section (left) - larger and more prominent
        time_frame = ctk.CTkFrame(info_frame, fg_color="transparent", width=120)
        time_frame.pack(side="left", fill="y")
        time_frame.pack_propagate(False)
        
        time_label = ctk.CTkLabel(
            time_frame,
            text=time_str,
            font=ctk.CTkFont(size=18, weight="bold"),
            anchor="w"
        )
        time_label.pack(pady=2)
        
        # Status section (center) - more prominent with attendance type
        status_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
        status_frame.pack(side="left", fill="both", expand=True, padx=20)
        
        # Get attendance type (default to CLOCK for existing records)
        attendance_type = record.get('attendance_type', 'CLOCK')
        
        # Build status display with location for CHECK records and late indicator
        is_late_clock_in = record.get('late', False) and attendance_type.upper() == 'CLOCK' and record['status'] == 'in'
        
        # Check for location text (for CHECK OUT records)
        location_text = ""
        if attendance_type.upper() == 'CHECK' and record['status'] == 'out':
            location_name = record.get('location_name', '')
            if location_name:
                # Truncate location name if too long - set to 30 characters
                if len(location_name) > 25:
                    location_name = location_name[:22] + "..."
                location_text = f"\n📍 {location_name}"
        
        if is_late_clock_in:
            # Create mixed-color status display for late clock-in
            mixed_record_frame = ctk.CTkFrame(status_frame, fg_color="transparent")
            mixed_record_frame.pack(expand=True)
            
            # Icon and LATE in red on first line
            first_line_frame = ctk.CTkFrame(mixed_record_frame, fg_color="transparent")
            first_line_frame.pack()
            
            icon_label = ctk.CTkLabel(
                first_line_frame,
                text=f"{status_icon} ",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=status_color
            )
            icon_label.pack(side="left")
            
            late_label = ctk.CTkLabel(
                first_line_frame,
                text="LATE ",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color="red"
            )
            late_label.pack(side="left")
            
            rest_label = ctk.CTkLabel(
                first_line_frame,
                text=f"{attendance_type} {status_text}",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=status_color
            )
            rest_label.pack(side="left")
            
            # Add location text if present
            if location_text:
                location_label = ctk.CTkLabel(
                    mixed_record_frame,
                    text=location_text,
                    font=ctk.CTkFont(size=14),
                    text_color=status_color
                )
                location_label.pack()
        else:
            # Normal single-color status display
            status_text_full = f"{status_icon} {attendance_type} {status_text}{location_text}"
            
            status_label = ctk.CTkLabel(
                status_frame,
                text=status_text_full,
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=status_color,
                anchor="center",
                justify="center"
            )
            status_label.pack(expand=True)
        
        # Method section (right) - more descriptive
        method_frame = ctk.CTkFrame(info_frame, fg_color="transparent", width=150)
        method_frame.pack(side="right", fill="y")
        method_frame.pack_propagate(False)
        
        method_label = ctk.CTkLabel(
            method_frame,
            text=method_text,
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray"),
            anchor="e"
        )
        method_label.pack(expand=True)
    
    
    def start_camera(self):
        """Start camera"""
        print("[CAMERA DEBUG] Attempting to start camera...")
        if self.camera_manager.start_camera():
            self.camera_active = True
            
            # Clear any cached recognition results for a fresh start
            self.face_recognition.clear_latest_results()
            print("[CAMERA DEBUG] Cleared cached recognition results for fresh start")
            
            # Set camera start time to ignore immediate recognition results
            self.camera_start_time = time.time()
            
            print("[CAMERA DEBUG] Camera started successfully")
        else:
            print("[CAMERA DEBUG] Failed to start camera")
            self.show_error_message("Cannot start camera")
    
    def stop_camera(self):
        """Stop camera"""
        print("[CAMERA DEBUG] Stopping camera...")
        self.camera_manager.stop_camera()
        self.camera_active = False
        
        # Clear cached recognition results to prevent old results from triggering dialogs
        self.face_recognition.clear_latest_results()
        print("[CAMERA DEBUG] Cleared cached recognition results")
        
        # Clear the camera display and show manual activation instructions
        self.camera_label.configure(
            image="", 
            text="📷 Camera Off\n\nPress + (Plus) or Numpad +\nto activate camera for recognition\n\nManual Recognition Mode"
        )
        print("[CAMERA DEBUG] Camera stopped")
    
    def toggle_camera(self):
        """Manual camera activation for recognition"""
        if not self.camera_active:
            # Start camera for recognition
            self.start_camera()
            if self.camera_active:
                self.show_success_message("📷 Camera activated for recognition\n\nPosition your face in front of the camera")
                print("[CAMERA DEBUG] Camera manually activated for recognition")
        else:
            # Camera is already active - this shouldn't happen in new workflow
            print("[CAMERA DEBUG] Camera already active - this is unexpected in manual mode")
            self.show_success_message("📷 Camera is already active")
    
    def export_daily_csv(self):
        """Show enhanced date selection dialog for Excel export"""
        self.show_export_date_selection_dialog()
    
    def show_export_date_selection_dialog(self):
        """Show comprehensive date selection dialog for Excel export with calendar picker"""
        from datetime import timedelta
        from tkinter import messagebox
        import calendar
        try:
            from tkcalendar import Calendar
        except ImportError:
            self.show_error_message("📅 tkcalendar module not found. Please install: pip install tkcalendar")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Excel Export - Calendar Date Selection")
        dialog.geometry("800x950")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (950 // 2)
        dialog.geometry(f"800x950+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='#f0f8ff')
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='#f0f8ff')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="📊 ATTENDANCE EXPORT",
            font=("Arial", 18, "bold"),
            bg='#f0f8ff',
            fg='#2E86C1'
        )
        title_label.pack(pady=10)
        
        # Export type selection
        export_frame = tk.LabelFrame(main_frame, text="Export Options", font=("Arial", 12, "bold"),
                                   bg='#f0f8ff', fg='#2E86C1', padx=10, pady=10)
        export_frame.pack(fill="x", pady=10)
        
        # Variable to track export type
        self.export_type = tk.StringVar(value="single_date")
        
        # Export type radio buttons
        tk.Radiobutton(export_frame, text="Single Date", variable=self.export_type, value="single_date",
                      font=("Arial", 11), bg='#f0f8ff', fg='black',
                      command=self.update_date_controls).pack(anchor="w", pady=2)
        
        tk.Radiobutton(export_frame, text="Date Range", variable=self.export_type, value="date_range",
                      font=("Arial", 11), bg='#f0f8ff', fg='black',
                      command=self.update_date_controls).pack(anchor="w", pady=2)
        
        tk.Radiobutton(export_frame, text="Whole Month", variable=self.export_type, value="whole_month",
                      font=("Arial", 11), bg='#f0f8ff', fg='black',
                      command=self.update_date_controls).pack(anchor="w", pady=2)
        
        tk.Radiobutton(export_frame, text="Whole Year", variable=self.export_type, value="whole_year",
                      font=("Arial", 11), bg='#f0f8ff', fg='black',
                      command=self.update_date_controls).pack(anchor="w", pady=2)
        
        # Date selection frame
        self.date_selection_frame = tk.LabelFrame(main_frame, text="Date Selection", 
                                                 font=("Arial", 12, "bold"),
                                                 bg='#f0f8ff', fg='#2E86C1', padx=10, pady=10)
        self.date_selection_frame.pack(fill="x", pady=10)
        
        # Initialize date variables
        today = datetime.now()
        self.start_year = tk.StringVar(value=str(today.year))
        self.start_month = tk.StringVar(value=str(today.month))
        self.start_day = tk.StringVar(value=str(today.day))
        self.end_year = tk.StringVar(value=str(today.year))
        self.end_month = tk.StringVar(value=str(today.month))
        self.end_day = tk.StringVar(value=str(today.day))
        
        # Create date controls (will be updated based on selection)
        self.create_date_controls()
        
        # Preview frame
        preview_frame = tk.LabelFrame(main_frame, text="Export Preview", 
                                    font=("Arial", 12, "bold"),
                                    bg='#f0f8ff', fg='#2E86C1', padx=10, pady=10)
        preview_frame.pack(fill="x", pady=10)
        
        self.preview_label = tk.Label(preview_frame, text="📊 Real-time preview will appear here",
                                    font=("Arial", 10), bg='#f0f8ff', fg='gray')
        self.preview_label.pack(pady=10)
        
        # Action buttons
        button_frame = tk.Frame(main_frame, bg='#f0f8ff')
        button_frame.pack(fill="x", pady=20)
        
        export_btn = tk.Button(button_frame, text="Export", 
                              font=("Arial", 12, "bold"),
                              command=self.execute_date_based_export,
                              bg='#2E86C1', fg='white', relief='raised', bd=2,
                              width=10)
        export_btn.pack(side="left", padx=10)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", 
                              font=("Arial", 12),
                              command=dialog.destroy,
                              bg='#ff6b6b', fg='white', relief='raised', bd=2,
                              width=10)
        cancel_btn.pack(side="right", padx=10)
        
        # Bind keyboard shortcuts
        dialog.bind('<Return>', lambda e: self.execute_date_based_export())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        dialog.focus_set()
        
        # Store dialog reference
        self.export_dialog = dialog
        
        # Initial update
        self.update_date_controls()
        
    def create_date_controls(self):
        """Create date input controls with calendar widgets and simplified selectors"""
        from tkcalendar import Calendar
        from tkinter import ttk
        import calendar
        
        # Clear existing controls
        for widget in self.date_selection_frame.winfo_children():
            widget.destroy()
        
        export_type = self.export_type.get()
        today = datetime.now()
        
        if export_type == "single_date":
            # Single date selection with calendar
            date_frame = tk.Frame(self.date_selection_frame, bg='#f0f8ff')
            date_frame.pack(pady=10)
            
            tk.Label(date_frame, text="📅 Select Date:", font=("Arial", 11, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').pack(pady=5)
            
            # Calendar widget for single date
            self.start_calendar = Calendar(
                date_frame,
                selectmode='day',
                year=today.year,
                month=today.month,
                day=today.day,
                background='white',
                foreground='black',
                bordercolor='#2E86C1',
                headersbackground='#2E86C1',
                headersforeground='white',
                selectbackground='#87CEEB',
                selectforeground='black',
                weekendbackground='#f0f8ff',
                weekendforeground='black',
                othermonthforeground='gray',
                othermonthbackground='#f5f5f5',
                font=('Arial', 10)
            )
            self.start_calendar.pack(pady=10)
            
            # Bind calendar selection for real-time updates
            self.start_calendar.bind("<<CalendarSelected>>", lambda e: self.update_export_preview())
            
            # Today button
            today_btn = tk.Button(date_frame, text="📅 Today", font=("Arial", 10, "bold"),
                                 command=self.set_today_calendar, bg='#87CEEB', fg='black',
                                 relief='raised', bd=2)
            today_btn.pack(pady=5)
            
        elif export_type == "date_range":
            # Date range selection with two calendars
            range_frame = tk.Frame(self.date_selection_frame, bg='#f0f8ff')
            range_frame.pack(pady=10)
            
            # Create a notebook-style layout for two calendars
            calendar_container = tk.Frame(range_frame, bg='#f0f8ff')
            calendar_container.pack(pady=5)
            
            # Start date calendar
            start_cal_frame = tk.LabelFrame(calendar_container, text="📅 From Date", 
                                          font=("Arial", 10, "bold"), bg='#f0f8ff', fg='#2E86C1')
            start_cal_frame.pack(side='left', padx=10, pady=5)
            
            self.start_calendar = Calendar(
                start_cal_frame,
                selectmode='day',
                year=today.year,
                month=today.month,
                day=1,  # Start of month
                background='white',
                foreground='black',
                bordercolor='#2E86C1',
                headersbackground='#2E86C1',
                headersforeground='white',
                selectbackground='#87CEEB',
                selectforeground='black',
                font=('Arial', 9)
            )
            self.start_calendar.pack(pady=5)
            
            # Bind for real-time updates
            self.start_calendar.bind("<<CalendarSelected>>", lambda e: self.update_export_preview())
            
            # End date calendar
            end_cal_frame = tk.LabelFrame(calendar_container, text="📅 To Date",
                                        font=("Arial", 10, "bold"), bg='#f0f8ff', fg='#2E86C1')
            end_cal_frame.pack(side='right', padx=10, pady=5)
            
            self.end_calendar = Calendar(
                end_cal_frame,
                selectmode='day',
                year=today.year,
                month=today.month,
                day=today.day,
                background='white',
                foreground='black',
                bordercolor='#2E86C1',
                headersbackground='#2E86C1',
                headersforeground='white',
                selectbackground='#87CEEB',
                selectforeground='black',
                font=('Arial', 9)
            )
            self.end_calendar.pack(pady=5)
            
            # Bind for real-time updates
            self.end_calendar.bind("<<CalendarSelected>>", lambda e: self.update_export_preview())
            
            # Quick range buttons
            quick_frame = tk.Frame(range_frame, bg='#f0f8ff')
            quick_frame.pack(pady=10)
            
            tk.Button(quick_frame, text="📅 Last 7 Days", command=lambda: self.set_quick_range_calendar(7),
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            tk.Button(quick_frame, text="📅 Last 30 Days", command=lambda: self.set_quick_range_calendar(30),
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            tk.Button(quick_frame, text="📅 This Month", command=self.set_current_month_calendar,
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            
        elif export_type == "whole_month":
            # Simplified month and year selection with dropdowns
            month_frame = tk.Frame(self.date_selection_frame, bg='#f0f8ff')
            month_frame.pack(pady=10)
            
            tk.Label(month_frame, text="🗓️ Select Month and Year:", font=("Arial", 11, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').pack(pady=5)
            
            # Month and Year dropdowns
            selector_frame = tk.Frame(month_frame, bg='#f0f8ff')
            selector_frame.pack(pady=10)
            
            # Month selection
            tk.Label(selector_frame, text="Month:", font=("Arial", 10, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=0, column=0, padx=5, pady=5, sticky='e')
            
            months = [calendar.month_name[i] for i in range(1, 13)]
            self.month_var = tk.StringVar(value=calendar.month_name[today.month])
            month_combo = ttk.Combobox(selector_frame, textvariable=self.month_var, values=months,
                                     state='readonly', font=("Arial", 10), width=12)
            month_combo.grid(row=0, column=1, padx=5, pady=5)
            month_combo.bind('<<ComboboxSelected>>', lambda e: self.update_export_preview())
            
            # Year selection
            tk.Label(selector_frame, text="Year:", font=("Arial", 10, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').grid(row=0, column=2, padx=5, pady=5, sticky='e')
            
            current_year = today.year
            years = [str(year) for year in range(1900, 2101)]
            self.year_var = tk.StringVar(value=str(today.year))
            year_combo = ttk.Combobox(selector_frame, textvariable=self.year_var, values=years,
                                    state='readonly', font=("Arial", 10), width=8, height=10)
            year_combo.grid(row=0, column=3, padx=5, pady=5)
            year_combo.bind('<<ComboboxSelected>>', lambda e: self.update_export_preview())
            
            # Quick month buttons
            quick_month_frame = tk.Frame(month_frame, bg='#f0f8ff')
            quick_month_frame.pack(pady=10)
            
            tk.Button(quick_month_frame, text="📅 This Month", command=self.set_current_month,
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            tk.Button(quick_month_frame, text="📅 Last Month", command=self.set_last_month,
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            
        elif export_type == "whole_year":
            # Simplified year selection with dropdown
            year_frame = tk.Frame(self.date_selection_frame, bg='#f0f8ff')
            year_frame.pack(pady=10)
            
            tk.Label(year_frame, text="📋 Select Year:", font=("Arial", 11, "bold"),
                    bg='#f0f8ff', fg='#2E86C1').pack(pady=5)
            
            # Year dropdown
            selector_frame = tk.Frame(year_frame, bg='#f0f8ff')
            selector_frame.pack(pady=10)
            
            current_year = today.year
            years = [str(year) for year in range(1900, 2101)]
            self.year_var = tk.StringVar(value=str(today.year))
            year_combo = ttk.Combobox(selector_frame, textvariable=self.year_var, values=years,
                                    state='readonly', font=("Arial", 12), width=15, height=10)
            year_combo.pack(pady=10)
            year_combo.bind('<<ComboboxSelected>>', lambda e: self.update_export_preview())
            
            # Quick year buttons
            quick_year_frame = tk.Frame(year_frame, bg='#f0f8ff')
            quick_year_frame.pack(pady=10)
            
            tk.Button(quick_year_frame, text="📅 This Year", command=self.set_current_year,
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
            tk.Button(quick_year_frame, text="📅 Last Year", command=self.set_last_year,
                     bg='#87CEEB', fg='black', font=("Arial", 9, "bold"), relief='raised', bd=2).pack(side="left", padx=5)
    
    def update_date_controls(self):
        """Update date controls based on export type selection"""
        self.create_date_controls()
        self.update_export_preview()
    
    def set_today_calendar(self):
        """Set calendar to today's date"""
        today = datetime.now()
        if hasattr(self, 'start_calendar'):
            self.start_calendar.selection_set(today.date())
        self.update_export_preview()
    
    def set_quick_range_calendar(self, days):
        """Set quick date range on calendars"""
        from datetime import timedelta
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        if hasattr(self, 'start_calendar'):
            self.start_calendar.selection_set(start_date.date())
        if hasattr(self, 'end_calendar'):
            self.end_calendar.selection_set(end_date.date())
        self.update_export_preview()
    
    def set_current_month_calendar(self):
        """Set calendars to current month range"""
        from datetime import timedelta
        today = datetime.now()
        start_date = datetime(today.year, today.month, 1)
        # Get last day of month
        if today.month == 12:
            end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
        
        if hasattr(self, 'start_calendar'):
            self.start_calendar.selection_set(start_date.date())
        if hasattr(self, 'end_calendar'):
            self.end_calendar.selection_set(end_date.date())
        self.update_export_preview()
    
    def set_current_month(self):
        """Set month/year selectors to current month"""
        import calendar
        today = datetime.now()
        if hasattr(self, 'month_var'):
            self.month_var.set(calendar.month_name[today.month])
        if hasattr(self, 'year_var'):
            self.year_var.set(str(today.year))
        self.update_export_preview()
    
    def set_last_month(self):
        """Set month/year selectors to last month"""
        import calendar
        from datetime import timedelta
        last_month_date = datetime.now().replace(day=1) - timedelta(days=1)
        if hasattr(self, 'month_var'):
            self.month_var.set(calendar.month_name[last_month_date.month])
        if hasattr(self, 'year_var'):
            self.year_var.set(str(last_month_date.year))
        self.update_export_preview()
    
    def set_current_year(self):
        """Set year selector to current year"""
        today = datetime.now()
        if hasattr(self, 'year_var'):
            self.year_var.set(str(today.year))
        self.update_export_preview()
    
    def set_last_year(self):
        """Set year selector to last year"""
        today = datetime.now()
        if hasattr(self, 'year_var'):
            self.year_var.set(str(today.year - 1))
        self.update_export_preview()
    
    def update_export_preview(self):
        """Update the export preview"""
        try:
            export_type = self.export_type.get()
            
            if export_type == "single_date":
                if hasattr(self, 'start_calendar'):
                    selected_date = self.start_calendar.selection_get()
                    date_str = selected_date.strftime("%Y-%m-%d")
                    preview_text = f"📅 Export Date: {date_str}"
                else:
                    preview_text = "📅 Select a date from the calendar"
                
            elif export_type == "date_range":
                if hasattr(self, 'start_calendar') and hasattr(self, 'end_calendar'):
                    start_date = self.start_calendar.selection_get()
                    end_date = self.end_calendar.selection_get()
                    start_str = start_date.strftime("%Y-%m-%d")
                    end_str = end_date.strftime("%Y-%m-%d")
                    preview_text = f"📆 Date Range: {start_str} to {end_str}"
                else:
                    preview_text = "📆 Select start and end dates from calendars"
                
            elif export_type == "whole_month":
                if hasattr(self, 'month_var') and hasattr(self, 'year_var'):
                    month_name = self.month_var.get()
                    year = self.year_var.get()
                    preview_text = f"🗓️ Whole Month: {month_name} {year}"
                else:
                    preview_text = "🗓️ Select month and year from dropdowns"
                
            elif export_type == "whole_year":
                if hasattr(self, 'year_var'):
                    year = self.year_var.get()
                    preview_text = f"📋 Whole Year: {year}"
                else:
                    preview_text = "📋 Select year from dropdown"
            
            # Get estimated record count
            record_count = self.get_estimated_record_count()
            preview_text += f"\n📊 Estimated Records: {record_count}"
            
            self.preview_label.configure(text=preview_text, fg='black')
            
        except Exception as e:
            self.preview_label.configure(text=f"⚠️ Preview Error: {str(e)}", fg='red')
    
    def get_estimated_record_count(self):
        """Get estimated record count for the selected date range"""
        try:
            export_type = self.export_type.get()
            
            if export_type == "single_date":
                if hasattr(self, 'start_calendar'):
                    selected_date = self.start_calendar.selection_get()
                    date_str = selected_date.strftime("%Y-%m-%d")
                    records = self.db.get_attendance_by_date_range(date_str, date_str)
                else:
                    return 0
                
            elif export_type == "date_range":
                if hasattr(self, 'start_calendar') and hasattr(self, 'end_calendar'):
                    start_date = self.start_calendar.selection_get().strftime("%Y-%m-%d")
                    end_date = self.end_calendar.selection_get().strftime("%Y-%m-%d")
                    records = self.db.get_attendance_by_date_range(start_date, end_date)
                else:
                    return 0
                
            elif export_type == "whole_month":
                if hasattr(self, 'month_var') and hasattr(self, 'year_var'):
                    import calendar
                    month_name = self.month_var.get()
                    year = int(self.year_var.get())
                    
                    # Convert month name to number
                    month_num = list(calendar.month_name).index(month_name)
                    
                    # Get first and last day of month
                    start_date = f"{year}-{month_num:02d}-01"
                    if month_num == 12:
                        end_date = f"{year + 1}-01-01"
                        end_date = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=1)
                        end_date = end_date.strftime("%Y-%m-%d")
                    else:
                        end_date = f"{year}-{month_num + 1:02d}-01"
                        end_date = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=1)
                        end_date = end_date.strftime("%Y-%m-%d")
                    
                    records = self.db.get_attendance_by_date_range(start_date, end_date)
                else:
                    return 0
                
            elif export_type == "whole_year":
                if hasattr(self, 'year_var'):
                    year = int(self.year_var.get())
                    start_date = f"{year}-01-01"
                    end_date = f"{year}-12-31"
                    records = self.db.get_attendance_by_date_range(start_date, end_date)
                else:
                    return 0
            
            return len(records) if records else 0
            
        except Exception as e:
            print(f"[EXPORT DEBUG] Error getting record count: {e}")
            return "Unknown"
    
    def execute_date_based_export(self):
        """Execute the export based on selected date criteria"""
        from tkinter import messagebox
        try:
            export_type = self.export_type.get()
            
            # Get the date range based on calendar selection
            if export_type == "single_date":
                if hasattr(self, 'start_calendar'):
                    selected_date = self.start_calendar.selection_get()
                    start_date = selected_date.strftime("%Y-%m-%d")
                    end_date = start_date
                    export_title = f"Single Day - {start_date}"
                else:
                    messagebox.showerror("Selection Error", "Please select a date from the calendar")
                    return
                
            elif export_type == "date_range":
                if hasattr(self, 'start_calendar') and hasattr(self, 'end_calendar'):
                    start_selected = self.start_calendar.selection_get()
                    end_selected = self.end_calendar.selection_get()
                    start_date = start_selected.strftime("%Y-%m-%d")
                    end_date = end_selected.strftime("%Y-%m-%d")
                    
                    # Ensure start date is before or equal to end date
                    if start_selected > end_selected:
                        messagebox.showerror("Date Error", "Start date must be before or equal to end date")
                        return
                        
                    export_title = f"Date Range - {start_date} to {end_date}"
                else:
                    messagebox.showerror("Selection Error", "Please select both start and end dates from the calendars")
                    return
                
            elif export_type == "whole_month":
                if hasattr(self, 'month_var') and hasattr(self, 'year_var'):
                    import calendar
                    month_name = self.month_var.get()
                    year = int(self.year_var.get())
                    
                    # Convert month name to number
                    month_num = list(calendar.month_name).index(month_name)
                    
                    start_date = f"{year}-{month_num:02d}-01"
                    if month_num == 12:
                        end_date = f"{year + 1}-01-01"
                        end_date = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=1)
                        end_date = end_date.strftime("%Y-%m-%d")
                    else:
                        end_date = f"{year}-{month_num + 1:02d}-01"
                        end_date = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=1)
                        end_date = end_date.strftime("%Y-%m-%d")
                    
                    export_title = f"Whole Month - {month_name} {year}"
                else:
                    messagebox.showerror("Selection Error", "Please select month and year from the dropdowns")
                    return
                
            elif export_type == "whole_year":
                if hasattr(self, 'year_var'):
                    year = int(self.year_var.get())
                    start_date = f"{year}-01-01"
                    end_date = f"{year}-12-31"
                    export_title = f"Whole Year - {year}"
                else:
                    messagebox.showerror("Selection Error", "Please select year from the dropdown")
                    return
            
            # Close the dialog first
            self.export_dialog.destroy()
            
            # Execute the export
            self.perform_date_range_export(start_date, end_date, export_title)
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error preparing export: {str(e)}")
    
    def perform_date_range_export(self, start_date, end_date, export_title):
        """Perform Excel export for the specified date range"""
        try:
            # Get attendance records for the date range
            records = self.db.get_attendance_by_date_range(start_date, end_date)
            
            if not records:
                self.show_error_message(f"📊 No attendance records found for {export_title}")
                return
            
            print(f"[EXCEL EXPORT] Found {len(records)} records for {export_title}")
            
            # Create exports directory if it doesn't exist
            exports_dir = "exports"
            if not os.path.exists(exports_dir):
                os.makedirs(exports_dir)
                
            # Generate filename with timestamp and date range
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_title = export_title.replace(" ", "_").replace("-", "_").replace("/", "_")
            filename = f"{exports_dir}/attendance_{safe_title}_{timestamp}.xlsx"
            
            # Create Excel workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = f"Attendance Export"
            
            # Define styles
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Late clock-in style (red background)
            late_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            late_font = Font(color="CC0000", bold=True)
            
            # On-time clock-in/clock-out style (green background)
            ontime_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            ontime_font = Font(color="006600", bold=True)
            
            # Regular style
            regular_alignment = Alignment(horizontal="left", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Excel Headers
            headers = [
                'Employee ID', 'Employee Name', 'Role', 'Date', 'Time', 'Status', 
                'Method', 'Attendance Type', 'Location Name', 'Address', 'Late Status'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Sort records by date, employee name, then by timestamp
            sorted_records = sorted(records, key=lambda x: (x['timestamp'][:10], x['name'], x['timestamp']))
            
            # Write data rows
            late_count = 0
            ontime_clockin_count = 0
            clockout_count = 0
            current_row = 2  # Start from row 2 (after header)
            previous_date = None
            
            for i, record in enumerate(sorted_records):
                # Parse timestamp
                record_datetime = datetime.strptime(record['timestamp'], "%Y-%m-%d %H:%M:%S")
                date_part = record_datetime.strftime("%Y-%m-%d")
                time_part = record_datetime.strftime("%H:%M:%S")
                
                # Check if this is a new date (different from previous record)
                if previous_date is not None and date_part != previous_date:
                    # Insert a blank row between different days
                    current_row += 1
                    # Add a subtle background color to the separator row for visual clarity
                    separator_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    for col in range(1, len(headers) + 1):
                        separator_cell = ws.cell(row=current_row, column=col, value="")
                        separator_cell.fill = separator_fill
                
                # Add date header for the first record of each day
                if previous_date != date_part:
                    current_row += 1
                    # Create a date header row
                    date_header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                    date_header_font = Font(bold=True, color="2E86C1")
                    date_header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Merge cells for date header
                    date_cell = ws.cell(row=current_row, column=1, value=f"📅 {date_part}")
                    date_cell.fill = date_header_fill
                    date_cell.font = date_header_font
                    date_cell.alignment = date_header_alignment
                    date_cell.border = border
                    
                    # Fill the rest of the merged row
                    for col in range(2, len(headers) + 1):
                        merge_cell = ws.cell(row=current_row, column=col, value="")
                        merge_cell.fill = date_header_fill
                        merge_cell.border = border
                    
                    # Merge the date header across all columns
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                
                # Update current row for actual data
                current_row += 1
                previous_date = date_part
                
                # Format status and attendance type
                attendance_type = record.get('attendance_type', 'check').upper()
                status_text = "IN" if record['status'] == 'in' else "OUT"
                
                # Format method
                method_map = {
                    'face_recognition': 'Face Recognition',
                    'qr_code': 'QR/Barcode Scan',
                    'manual': 'Manual Entry'
                }
                method = method_map.get(record['method'], record['method'])
                
                # Get location data (only for CHECK OUT records)
                location_name = record.get('location_name', '') if record['status'] == 'out' else ''
                address = record.get('address', '') if record['status'] == 'out' else ''
                
                # Check if this is a late record
                is_late = record.get('late', False)
                late_status = 'LATE' if is_late else 'ON-TIME'
                
                # Get employee role
                employee_data = self.db.get_employee(record['employee_id'])
                employee_role = employee_data.get('role', 'Staff') if employee_data else 'Staff'
                role_display = f"🛡️ {employee_role}" if employee_role == 'Security' else f"👥 {employee_role}"
                
                # Write row data
                row_data = [
                    record['employee_id'],
                    record['name'],
                    role_display,
                    date_part,
                    time_part,
                    status_text,
                    method,
                    attendance_type,
                    location_name,
                    address,
                    late_status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.alignment = regular_alignment
                    cell.border = border
                    
                    # Highlight clock records based on status and late flag
                    if attendance_type == 'CLOCK':
                        if record['status'] == 'in':
                            if is_late:
                                # Late clock-in - red highlighting
                                cell.fill = late_fill
                                if col == 11:  # Late Status column (shifted by 1 due to Role column)
                                    cell.font = late_font
                                late_count += 1
                            else:
                                # On-time clock-in - green highlighting
                                cell.fill = ontime_fill
                                if col == 11:  # Late Status column (shifted by 1 due to Role column)
                                    cell.font = ontime_font
                                ontime_clockin_count += 1
                        elif record['status'] == 'out':
                            # Clock-out - green highlighting (always on time due to restrictions)
                            cell.fill = ontime_fill
                            clockout_count += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
            # Show success message with file location and statistics
            success_msg = f"📊 Exported {len(records)} records\n📅 {export_title}\n💾 Saved to: {filename}"
            
            if late_count > 0:
                success_msg += f"\n🔴 {late_count} late clock-ins highlighted"
            if ontime_clockin_count > 0:
                success_msg += f"\n🟢 {ontime_clockin_count} on-time clock-ins highlighted"
            if clockout_count > 0:
                success_msg += f"\n🟢 {clockout_count} clock-outs highlighted"
            
            self.show_success_message(success_msg)
            print(f"[EXCEL EXPORT] Successfully exported to: {filename}")
            print(f"[EXCEL EXPORT] Date range: {start_date} to {end_date}")
            if late_count > 0:
                print(f"[EXCEL EXPORT] {late_count} late clock-ins highlighted in red")
            if ontime_clockin_count > 0:
                print(f"[EXCEL EXPORT] {ontime_clockin_count} on-time clock-ins highlighted in green")
            if clockout_count > 0:
                print(f"[EXCEL EXPORT] {clockout_count} clock-outs highlighted in green")
            
        except Exception as e:
            print(f"[EXCEL EXPORT] Error exporting Excel: {e}")
            self.show_error_message(f"📊 Excel export failed: {str(e)}")
    
    def show_excel_export_confirmation(self, date, record_count):
        """Show confirmation dialog for Excel export"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Excel Export Confirmation")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog on screen
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (350 // 2)
        dialog.geometry(f"450x350+{x}+{y}")
        
        # Configure dialog background
        dialog.configure(bg='#2E86C1')
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='#2E86C1')
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Export icon
        tk.Label(main_frame, text="�", font=("Arial", 32), 
                bg='#2E86C1', fg='white').pack(pady=10)
        
        # Confirmation message
        tk.Label(main_frame, text="Export", 
                font=("Arial", 14, "bold"), bg='#2E86C1', fg='white').pack()
        
        tk.Label(main_frame, text=f"Date: {date}", 
                font=("Arial", 12), bg='#2E86C1', fg='white').pack(pady=5)
        
        tk.Label(main_frame, text=f"Records: {record_count}", 
                font=("Arial", 12), bg='#2E86C1', fg='white').pack(pady=2)
        
        tk.Label(main_frame, text="• Late clock-ins highlighted in red", 
                font=("Arial", 10), bg='#2E86C1', fg='lightblue').pack(pady=2)
        
        tk.Label(main_frame, text="• Professional Excel formatting", 
                font=("Arial", 10), bg='#2E86C1', fg='lightblue').pack(pady=2)
        
        tk.Label(main_frame, text="Proceed with Excel export?", 
                font=("Arial", 11, "bold"), bg='#2E86C1', fg='white').pack(pady=15)
        
        # Bind keys for confirmation
        def confirm_export():
            dialog.destroy()
            self.perform_excel_export()
            
        def cancel_export():
            dialog.destroy()
            self.show_success_message("Excel export cancelled")
        
        dialog.bind('<Return>', lambda e: confirm_export())       # Enter key
        dialog.bind('<KP_Enter>', lambda e: confirm_export())     # Numpad Enter
        dialog.bind('<KP_Multiply>', lambda e: confirm_export())  # Numpad *
        dialog.bind('<asterisk>', lambda e: confirm_export())     # Regular *
        dialog.bind('<KP_Subtract>', lambda e: cancel_export())   # Numpad -
        dialog.bind('<minus>', lambda e: cancel_export())         # Regular -
        dialog.bind('<Escape>', lambda e: cancel_export())        # Escape key
        
        dialog.focus_set()
        
        # Auto-cancel after 15 seconds
        dialog.after(15000, lambda: dialog.destroy() if dialog.winfo_exists() else None)
    
    def perform_excel_export(self):
        """Actually perform the Excel export with late highlighting"""
        try:
            # Get today's date
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # Get today's attendance records
            attendance_records = self.db.get_attendance_today()
            
            # Filter to ensure only today's records (extra safety)
            today_records = []
            for record in attendance_records:
                record_date = record['timestamp'][:10]  # Extract YYYY-MM-DD part
                if record_date == current_date:
                    today_records.append(record)
            
            print(f"[EXCEL EXPORT] Found {len(today_records)} records for {current_date}")
            
            if not today_records:
                self.show_error_message("No attendance records found for today")
                return
            
            # Create exports directory if it doesn't exist
            exports_dir = "exports"
            if not os.path.exists(exports_dir):
                os.makedirs(exports_dir)
                
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{exports_dir}/attendance_{current_date}_{timestamp}.xlsx"
            
            # Create Excel workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = f"Attendance {current_date}"
            
            # Define styles
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Late clock-in style (red background)
            late_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            late_font = Font(color="CC0000", bold=True)
            
            # On-time clock-in style (green background)
            ontime_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            ontime_font = Font(color="006600", bold=True)

            # On-time clock-out style (yellow background)
            ontime_fill_clockout = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
            
            # Regular style
            regular_alignment = Alignment(horizontal="left", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Excel Headers
            headers = [
                'Employee ID', 'Employee Name', 'Date', 'Time', 'Status', 
                'Method', 'Attendance Type', 'Location Name', 'Address', 'Late Status'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Sort records by employee name, then by timestamp ascending
            sorted_records = sorted(today_records, key=lambda x: (x['name'], x['timestamp']))
            
            # Write data rows
            late_count = 0
            for row, record in enumerate(sorted_records, 2):  # Start from row 2 (after header)
                # Parse timestamp
                record_datetime = datetime.strptime(record['timestamp'], "%Y-%m-%d %H:%M:%S")
                date_part = record_datetime.strftime("%Y-%m-%d")
                time_part = record_datetime.strftime("%H:%M:%S")
                
                # Format status and attendance type
                attendance_type = record.get('attendance_type', 'check').upper()
                if attendance_type == 'CLOCK':
                    status = "CLOCK IN" if record['status'] == 'in' else "CLOCK OUT"
                else:  # CHECK
                    status = "CHECK IN" if record['status'] == 'in' else "CHECK OUT"
                
                # Format method
                method_map = {
                    'face_recognition': 'Face Recognition',
                    'qr_code': 'QR/Barcode Scan',
                    'manual': 'Manual Entry'
                }
                method = method_map.get(record['method'], record['method'])
                
                # Get location data (only for CHECK OUT records)
                location_name = record.get('location_name', '') if record['status'] == 'out' else ''
                address = record.get('address', '') if record['status'] == 'out' else ''
                
                # Check if this is a late record
                is_late = record.get('late', False)
                late_status = 'LATE' if is_late else 'ON-TIME'
                
                row_data = [
                    record['employee_id'],
                    record['name'],
                    date_part,
                    time_part,
                    status,
                    method,
                    attendance_type,
                    location_name,
                    address,
                    late_status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = regular_alignment
                    cell.border = border
                    
                    # Highlight clock records based on status and late flag
                    if attendance_type == 'CLOCK':
                        if record['status'] == 'in':
                            if is_late:
                                # Late clock-in - red highlighting
                                cell.fill = late_fill
                                if col == 10:  # Late Status column
                                    cell.font = late_font
                                late_count += 1
                            else:
                                # On-time clock-in - green highlighting
                                cell.fill = ontime_fill
                                if col == 10:  # Late Status column
                                    cell.font = ontime_font
                        elif record['status'] == 'out':
                            # Clock-out - green highlighting (always on time due to restrictions)
                            cell.fill = ontime_fill_clockout
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
            # Count actual late and on-time clock records for reporting
            actual_late_count = sum(1 for r in today_records 
                                   if r.get('late', False) and 
                                      r.get('attendance_type') == 'clock' and 
                                      r['status'] == 'in')
            
            # Count on-time clock-ins and all clock-outs (green highlighting)
            actual_ontime_count = sum(1 for r in today_records 
                                     if ((not r.get('late', False) and r['status'] == 'in') or r['status'] == 'out') and
                                        r.get('attendance_type') == 'clock')
            
            # Count clock-outs separately for detailed reporting
            clockout_count = sum(1 for r in today_records 
                               if r.get('attendance_type') == 'clock' and r['status'] == 'out')
            
            # Show success message with file location
            success_msg = f"Exported {len(today_records)} records\nSaved to: {filename}"
            if actual_late_count > 0:
                success_msg += f"\n🔴 {actual_late_count} late clock-ins highlighted"
            # Count on-time clock-ins separately for clearer messaging
            ontime_clockin_count = sum(1 for r in today_records 
                                     if not r.get('late', False) and 
                                        r.get('attendance_type') == 'clock' and 
                                        r['status'] == 'in')
            
            if ontime_clockin_count > 0:
                success_msg += f"\n🟢 {ontime_clockin_count} on-time clock-ins highlighted"
            if clockout_count > 0:
                success_msg += f"\n🟢 {clockout_count} clock-outs highlighted"
            
            self.show_success_message(success_msg)
            print(f"[EXCEL EXPORT] Successfully exported to: {filename}")
            if actual_late_count > 0:
                print(f"[EXCEL EXPORT] {actual_late_count} late clock-ins highlighted in red")
            if ontime_clockin_count > 0:
                print(f"[EXCEL EXPORT] {ontime_clockin_count} on-time clock-ins highlighted in green")
            if clockout_count > 0:
                print(f"[EXCEL EXPORT] {clockout_count} clock-outs highlighted in green")
            
        except Exception as e:
            print(f"[EXCEL EXPORT] Error exporting Excel: {e}")
            self.show_error_message(f"Excel export failed: {str(e)}")
    
    def update_loop(self):
        """Main update loop with comprehensive error handling"""
        try:
            # Update time
            current_time = datetime.now().strftime("%A, %B %d, %Y - %H:%M:%S")
            self.time_label.configure(text=current_time)
            
            # Update attendance history every 30 seconds
            current_timestamp = time.time()
            if current_timestamp - self.last_history_update > 30:
                self.update_attendance_history()
                self.last_history_update = current_timestamp
            
            # Check focus every 10 seconds to maintain keyboard control
            if not hasattr(self, 'last_focus_check'):
                self.last_focus_check = 0
            if current_timestamp - self.last_focus_check > 10:
                self.maintain_focus()
                self.last_focus_check = current_timestamp
            
            # Process camera with error handling
            if self.camera_active:
                try:
                    self.process_camera()
                except Exception as cam_e:
                    print(f"[UPDATE ERROR] Camera processing failed: {cam_e}")
                    # Force garbage collection on camera errors
                    import gc
                    collected = gc.collect()
                    print(f"[GC ERROR] Collected {collected} objects after camera error")
            
        except Exception as e:
            print(f"[UPDATE ERROR] Exception in main update loop: {e}")
            import traceback
            traceback.print_exc()
            # Force garbage collection on any update error
            import gc
            collected = gc.collect()
            print(f"[GC ERROR] Collected {collected} objects after update error")
        
        finally:
            # Always schedule next update (16ms ≈ 60 FPS for smoother video)
            try:
                self.root.after(16, self.update_loop)
            except Exception as e:
                print(f"[SCHEDULE ERROR] Failed to schedule next update: {e}")
                # Try with longer delay as fallback
                try:
                    self.root.after(50, self.update_loop)
                except:
                    print("[CRITICAL ERROR] Cannot schedule updates - application may freeze")
    
    def process_camera(self):
        """Process camera feed with error handling and restart capability"""
        try:
            # Skip camera processing if camera is not active (manual mode)
            if not self.camera_active:
                return  # Camera is OFF, no processing needed
                
            # Check if main camera is paused (e.g., during registration)
            if hasattr(self, 'main_camera_paused') and self.main_camera_paused:
                # Debug output only once per pause to avoid spam
                if not hasattr(self, '_pause_debug_shown') or not self._pause_debug_shown:
                    print("[CAMERA DEBUG] Main camera processing paused for registration")
                    # Show paused status on camera display
                    try:
                        self.camera_label.configure(image="", text="📷 Camera Paused\n(Registration Active)")
                    except:
                        pass
                    self._pause_debug_shown = True
                return  # Skip camera processing when paused
            
            # Reset debug flag when not paused
            if hasattr(self, '_pause_debug_shown'):
                if self._pause_debug_shown:
                    print("[CAMERA DEBUG] Main camera processing resumed")
                self._pause_debug_shown = False
            
            # Read frame with error checking
            frame = self.camera_manager.read_frame()
            if frame is None:
                # Try to restart camera if read fails multiple times
                if not hasattr(self, 'camera_fail_count'):
                    self.camera_fail_count = 0
                self.camera_fail_count += 1
                
                if self.camera_fail_count > 30:  # 30 failed reads in a row
                    print("[CAMERA ERROR] Multiple camera read failures, attempting restart...")
                    self.restart_camera()
                    self.camera_fail_count = 0
                return
            
            # Reset failure count on successful read
            if hasattr(self, 'camera_fail_count'):
                self.camera_fail_count = 0
            
            # Resize for display - smaller to fit in the compact camera frame
            height, width = frame.shape[:2]
            display_width = 450  # Reduced from 600 to fit better
            display_height = int(height * display_width / width)
            # Limit height to fit in camera frame
            max_height = 350
            if display_height > max_height:
                display_height = max_height
                display_width = int(width * display_height / height)
            display_frame = cv2.resize(frame, (display_width, display_height))
            
            # Choose face detection method based on configuration
            if self.use_ultra_light_detection:
                # Use Ultra Light Face Detection for maximum performance
                recognized_faces = self._process_ultra_light_detection(frame, display_frame, display_width, display_height, width, height)
            else:
                # Use traditional DeepFace processing
                recognized_faces = self._process_deepface_detection(frame, display_frame, display_width, display_height, width, height)
            
            if recognized_faces:
                print(f"[FACE DEBUG] Detected {len(recognized_faces)} face(s)")
        
        except Exception as e:
            print(f"[CAMERA ERROR] Exception in camera processing: {e}")
            # Force garbage collection on camera errors
            import gc
            collected = gc.collect()
            print(f"[GC ERROR] Collected {collected} objects after camera error")
            return
            
        # Process recognized faces for attendance
        current_time = time.time()  # Add timing for recognition logic
        
        # Ignore recognition results for first 2 seconds after camera start to prevent immediate dialog
        if hasattr(self, 'camera_start_time') and (current_time - self.camera_start_time) < 2.0:
            print(f"[CAMERA DEBUG] Ignoring recognition results for {2.0 - (current_time - self.camera_start_time):.1f}s after camera start")
            return
        
        for face in recognized_faces:
            if face['employee_id']:
                # Check if personal scanning is active
                if hasattr(self, 'personal_scanning_active') and self.personal_scanning_active:
                    # Auto-fill employee ID in personal dialog
                    if hasattr(self, 'personal_id_entry'):
                        self.personal_id_entry.delete(0, tk.END)
                        self.personal_id_entry.insert(0, face['employee_id'])
                        print(f"[PERSONAL DEBUG] Face recognition auto-filled: {face['employee_id']}")
                    return
                
                # Check if group scanning is active
                if hasattr(self, 'group_scanning_active') and self.group_scanning_active:
                    # Auto-fill employee ID in group dialog
                    if hasattr(self, 'group_current_id'):
                        self.group_current_id.delete(0, tk.END)
                        self.group_current_id.insert(0, face['employee_id'])
                        print(f"[GROUP DEBUG] Face recognition auto-filled: {face['employee_id']}")
                    return
                
                print(f"[FACE DEBUG] Recognized employee: {face['name']} ({face['employee_id']})")
                employee = self.db.get_employee(face['employee_id'])
                
                # Show confirmation dialog before processing attendance
                self.show_face_recognition_confirmation(face, employee)
                return
            else:
                print(f"[FACE DEBUG] Detected unknown face")
        
        # Try QR code scanning
        detected_codes = self.barcode_scanner.scan_frame(frame)
        if detected_codes and len(detected_codes) > 0:
            # Get the first detected code and extract the employee ID
            first_code = detected_codes[0]
            employee_id = first_code.get('employee_id')
            
            if employee_id:  # Valid employee ID found
                # Check if personal scanning is active
                if hasattr(self, 'personal_scanning_active') and self.personal_scanning_active:
                    # Auto-fill employee ID in personal dialog
                    if hasattr(self, 'personal_id_entry'):
                        self.personal_id_entry.delete(0, tk.END)
                        self.personal_id_entry.insert(0, employee_id)
                        print(f"[PERSONAL DEBUG] QR scan auto-filled: {employee_id}")
                    return
                
                # Check if group scanning is active
                if hasattr(self, 'group_scanning_active') and self.group_scanning_active:
                    # Auto-fill employee ID in group dialog
                    if hasattr(self, 'group_current_id'):
                        self.group_current_id.delete(0, tk.END)
                        self.group_current_id.insert(0, employee_id)
                        print(f"[GROUP DEBUG] QR scan auto-filled: {employee_id}")
                    return
                
                print(f"[QR SCAN DEBUG] Detected QR/Barcode: '{employee_id}' from data: '{first_code.get('data')}'")
                employee = self.db.get_employee(employee_id)
                if employee:
                    print(f"[QR SCAN DEBUG] Found employee: {employee['name']} ({employee['employee_id']})")
                    
                    # Show confirmation dialog before processing attendance
                    self.show_qr_recognition_confirmation(first_code, employee)
                else:
                    print(f"[QR SCAN DEBUG] Employee not found in database")
                    self.show_error_message(f"✗ Employee {employee_id} not found")
            else:
                print(f"[QR SCAN DEBUG] Invalid QR code data: '{first_code.get('data')}'")
                self.show_error_message(f"✗ Invalid QR code format")
            return
        else:
            # Only print this occasionally to avoid spam
            if hasattr(self, 'debug_counter'):
                self.debug_counter += 1
            else:
                self.debug_counter = 0
            
            if self.debug_counter % 300 == 0:  # Print every 300 frames (about every 5 seconds at 60 FPS)
                print(f"[QR SCAN DEBUG] No QR/Barcode detected in frame {self.debug_counter}")
        
        # Draw face boxes if any detected
        if recognized_faces:
            display_frame = self.face_recognition.draw_face_boxes_from_results(display_frame, recognized_faces)
        else:
            # No faces detected - show positioning guidance
            self._display_distance_feedback(display_frame, "Please position your face in the camera view", False)
        
        # Ensure display_frame is valid before color conversion
        if display_frame is None:
            print("[CAMERA ERROR] Display frame is None")
            return
        
        # Convert and display frame with error handling
        try:
            frame_rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
            frame_pil = Image.fromarray(frame_rgb)
            frame_tk = ImageTk.PhotoImage(frame_pil)
            
            self.camera_label.configure(image=frame_tk, text="")
            self.camera_label.image = frame_tk
        except Exception as e:
            print(f"[CAMERA ERROR] Failed to display frame: {e}")
            # Show error text instead of crashing
            try:
                self.camera_label.configure(image="", text="📷 Camera Error\nRestarting...")
            except:
                pass  # Ignore secondary errors
    
    def quit_app(self):
        """Quit application with confirmation"""
        print("[QUIT DEBUG] Quit requested - showing confirmation")
        
        # Create a full-screen overlay for confirmation
        self.quit_overlay = ctk.CTkToplevel(self.root)
        self.quit_overlay.title("Confirm Exit")
        self.quit_overlay.attributes('-fullscreen', True)
        self.quit_overlay.attributes('-topmost', True)
        self.quit_overlay.configure(fg_color=("gray10", "gray10"))  # Semi-transparent dark
        self.quit_overlay.transient(self.root)
        self.quit_overlay.grab_set()
        
        # Center the confirmation message
        confirm_frame = ctk.CTkFrame(
            self.quit_overlay,
            width=800,
            height=300,
            corner_radius=20,
            fg_color=("red", "darkred"),
            border_width=5,
            border_color="white"
        )
        confirm_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        # Confirmation message
        warning_label = ctk.CTkLabel(
            confirm_frame,
            text="⚠️",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color="white"
        )
        warning_label.pack(pady=40)
        
        instruction_label = ctk.CTkLabel(
            confirm_frame,
            text="Press - again to EXIT\nPress Enter to CANCEL",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="white"
        )
        instruction_label.pack(pady=20)
        
        # Force UI update
        self.root.update()
        print("[QUIT DEBUG] Confirmation overlay displayed")
        
        # Temporarily change key bindings for confirmation
        self.root.unbind('<KP_Subtract>')
        self.root.unbind('<minus>')
        self.root.unbind('<Return>')
        self.root.unbind('<KP_Enter>')
        
        self.quit_overlay.bind('<KP_Subtract>', lambda e: self.confirm_quit())     # Numpad - to confirm
        self.quit_overlay.bind('<minus>', lambda e: self.confirm_quit())           # Regular - to confirm
        self.quit_overlay.bind('<Return>', lambda e: self.cancel_quit())          # Enter to cancel
        self.quit_overlay.bind('<KP_Enter>', lambda e: self.cancel_quit())        # Numpad Enter to cancel
        self.quit_overlay.bind('<Escape>', lambda e: self.cancel_quit())          # ESC to cancel
        
        # Focus on overlay to capture keys
        self.quit_overlay.focus_set()
        
        # Auto-cancel after 5 seconds
        self.root.after(5000, self.cancel_quit)
    
    def confirm_quit(self):
        """Actually quit the application"""
        print("[QUIT DEBUG] Quit confirmed - exiting application")
        # Close the confirmation overlay
        if hasattr(self, 'quit_overlay') and self.quit_overlay:
            self.quit_overlay.destroy()
        # Stop background face processing
        if hasattr(self, 'face_recognition') and self.face_recognition:
            self.face_recognition.stop_background_processing()
        # Close MongoDB connection
        if hasattr(self, 'db') and self.db:
            self.db.close_connection()
        if self.camera_active:
            self.camera_manager.stop_camera()
        self.root.quit()
        self.root.destroy()
    
    def cancel_quit(self):
        """Cancel the quit operation"""
        print("[QUIT DEBUG] Quit canceled - returning to normal operation")
        # Close the confirmation overlay
        if hasattr(self, 'quit_overlay') and self.quit_overlay:
            self.quit_overlay.destroy()
            self.quit_overlay = None
        # Restore original key bindings
        self.setup_keyboard_shortcuts()
    
    def restart_camera(self):
        """Restart camera on failure"""
        try:
            print("[CAMERA RESTART] Stopping camera...")
            self.camera_manager.stop_camera()
            time.sleep(1)  # Brief pause
            
            print("[CAMERA RESTART] Reinitializing camera...")
            self.camera_manager = CameraManager()
            
            if self.camera_active:
                print("[CAMERA RESTART] Starting camera...")
                self.camera_manager.start_camera()
                print("[CAMERA RESTART] Camera restarted successfully")
            
        except Exception as e:
            print(f"[CAMERA RESTART] Failed to restart camera: {e}")
    
    def _process_ultra_light_detection(self, frame, display_frame, display_width, display_height, width, height):
        """Process frame using Ultra Light Face Detection for maximum performance"""
        recognized_faces = []
        
        try:
            # Detect faces using ultra light detector
            detected_faces = self.ultra_light_detector.detect_faces_for_attendance(frame)
            
            if detected_faces:
                # Get the best face for attendance
                best_face = self.ultra_light_detector.get_best_face(detected_faces)
                
                # Check distance for the best face and provide feedback
                if best_face:
                    x1, y1, x2, y2 = best_face['bbox']
                    face_coords = (x1, y1, x2 - x1, y2 - y1)  # Convert to (x, y, w, h) format
                    is_good_distance, feedback_message = self._check_face_distance_and_provide_feedback(frame, face_coords)
                    
                    # Display distance feedback on the frame
                    self._display_distance_feedback(display_frame, feedback_message, is_good_distance)
                
                # Draw all detected faces on display frame
                display_faces_with_labels = []
                
                for face_data in detected_faces:
                    x1, y1, x2, y2 = face_data['bbox']
                    confidence = face_data['confidence']
                    
                    # Scale coordinates to display frame
                    display_x1 = int(x1 * display_width / width)
                    display_y1 = int(y1 * display_height / height)
                    display_x2 = int(x2 * display_width / width)
                    display_y2 = int(y2 * display_height / height)
                    
                    # Extract face region for recognition
                    face_region = frame[y1:y2, x1:x2]
                    
                    # Perform face recognition on the detected face
                    employee_id = None
                    employee_name = "Unknown"
                    recognition_confidence = 0.0
                    
                    print(f"[ULTRA DEBUG] Face region size: {face_region.shape if face_region.size > 0 else 'empty'}")
                    print(f"[ULTRA DEBUG] Bbox coordinates: ({x1}, {y1}, {x2}, {y2})")
                    
                    if face_region.size > 0 and face_region.shape[0] > 20 and face_region.shape[1] > 20:  # Valid face region with minimum size
                        try:
                            print(f"[ULTRA DEBUG] Calling DeepFace recognition on face region...")
                            print(f"[ULTRA DEBUG] Face region shape: {face_region.shape}")
                            
                            # Use DeepFace recognition on the detected face region
                            recognized_id, rec_conf = self.face_recognition.recognize_face(face_region)
                            print(f"[ULTRA DEBUG] DeepFace returned: ID={recognized_id}, confidence={rec_conf}")
                            
                            # If direct recognition fails, try with some padding around the face
                            if not recognized_id:
                                print(f"[ULTRA DEBUG] Direct recognition failed, trying with padded region...")
                                
                                # Add padding around the detected face
                                padding = 30
                                padded_x1 = max(0, x1 - padding)
                                padded_y1 = max(0, y1 - padding) 
                                padded_x2 = min(frame.shape[1], x2 + padding)
                                padded_y2 = min(frame.shape[0], y2 + padding)
                                
                                padded_face_region = frame[padded_y1:padded_y2, padded_x1:padded_x2]
                                print(f"[ULTRA DEBUG] Padded region shape: {padded_face_region.shape}")
                                
                                if padded_face_region.size > 0:
                                    recognized_id, rec_conf = self.face_recognition.recognize_face(padded_face_region)
                                    print(f"[ULTRA DEBUG] Padded recognition returned: ID={recognized_id}, confidence={rec_conf}")
                            
                            if recognized_id:
                                employee_id = recognized_id
                                recognition_confidence = rec_conf
                                # Get employee name from known faces
                                if recognized_id in self.face_recognition.known_faces:
                                    employee_name = self.face_recognition.known_faces[recognized_id].get('name', recognized_id)
                                print(f"[ULTRA RECOGNITION] Face recognized: {employee_name} ({employee_id}) confidence: {rec_conf:.2f}")
                            else:
                                print(f"[ULTRA RECOGNITION] Face detected but not recognized (detection conf: {confidence:.2f})")
                        except Exception as rec_error:
                            print(f"[ULTRA RECOGNITION] Recognition error: {rec_error}")
                    else:
                        print(f"[ULTRA DEBUG] Face region too small or invalid for recognition")
                    
                    # Create face data in format expected by main drawing system
                    is_best = best_face and face_data['id'] == best_face['id']
                    
                    if employee_id:
                        # Recognized face
                        label = f"{employee_name} ({recognition_confidence:.2f})"
                    else:
                        # Detected but unrecognized face
                        label = f"Face Detected ({confidence:.2f})"
                    
                    if is_best:
                        label += " [BEST]"
                    
                    face_result = {
                        'name': employee_name if employee_id else label,
                        'employee_id': employee_id,  # Now includes actual recognition results
                        'confidence': recognition_confidence if employee_id else confidence,
                        'position': (display_x1, display_y1, display_x2 - display_x1, display_y2 - display_y1),
                        'is_best': is_best
                    }
                    display_faces_with_labels.append(face_result)
                    
                    # Add to recognized_faces so main system can draw them
                    recognized_faces.append(face_result)
                
                # Don't draw faces here - let the main system handle it
                # display_frame = self._draw_ultra_light_faces(display_frame, display_faces_with_labels)
                
                # For attendance processing, we could integrate with recognition here
                # For now, ultra light detection is mainly for showing real-time face detection
                # You could add face recognition integration here if needed
            else:
                # No faces detected by Ultra Light detector - show positioning guidance
                self._display_distance_feedback(display_frame, "Please position your face in the camera view", False)
                
        except Exception as e:
            print(f"[ULTRA LIGHT ERROR] Detection processing failed: {e}")
        
        return recognized_faces
    
    def _process_deepface_detection(self, frame, display_frame, display_width, display_height, width, height):
        """Process frame using traditional DeepFace detection"""
        recognized_faces = []
        
        try:
            # Submit frame for background DeepFace processing (non-blocking)
            self.face_recognition.submit_frame_for_processing(frame)
            
            # Get latest recognition results from background thread
            face_results = self.face_recognition.get_latest_results()
            
            if face_results:
                # Check distance for the first detected face and provide feedback
                best_face = None
                best_confidence = 0.0
                
                # Find the face with highest confidence for distance checking
                for face in face_results:
                    if 'position' in face and face.get('confidence', 0.0) > best_confidence:
                        best_face = face
                        best_confidence = face.get('confidence', 0.0)
                
                if best_face:
                    x, y, w, h = best_face['position']
                    face_coords = (x, y, w, h)
                    is_good_distance, feedback_message = self._check_face_distance_and_provide_feedback(frame, face_coords)
                    
                    # Display distance feedback on the frame
                    self._display_distance_feedback(display_frame, feedback_message, is_good_distance)
                
                # Convert coordinates to display frame and draw faces
                display_faces_with_labels = []
                
                for face in face_results:
                    if 'position' in face:
                        face_confidence = face.get('confidence', 0.0)
                        x, y, w, h = face['position']
                        # Scale coordinates to display frame
                        display_x = int(x * display_width / width)
                        display_y = int(y * display_height / height)
                        display_w = int(w * display_width / width)
                        display_h = int(h * display_height / height)
                        
                        # Check confidence and handle accordingly
                        CONFIDENCE_THRESHOLD = 0.6  # Define threshold
                        if face_confidence < CONFIDENCE_THRESHOLD:
                            # Show low confidence faces with special label
                            print(f"[CONFIDENCE NOTICE] Low confidence detection: {face_confidence:.2f} < {CONFIDENCE_THRESHOLD}")
                            display_result = {
                                'name': 'Low Confidence',
                                'employee_id': None,
                                'confidence': face_confidence,
                                'position': (display_x, display_y, display_w, display_h)
                            }
                            display_faces_with_labels.append(display_result)
                            # Don't add to recognized_faces (no attendance processing)
                        else:
                            # High confidence face - process normally
                            display_result = {
                                'name': face.get('name'),
                                'employee_id': face.get('employee_id'),
                                'confidence': face.get('confidence', 0.0),
                                'position': (display_x, display_y, display_w, display_h)
                            }
                            display_faces_with_labels.append(display_result)
                            
                            # Store recognized faces for attendance processing
                            if face['employee_id']:
                                recognized_faces.append(face)
                                print(f"[FACE DEBUG] Recognized: {face['name']} (ID: {face['employee_id']}, confidence: {face['confidence']:.2f})")
                
                # Draw faces with proper labels using pure DeepFace results
                display_frame = self.face_recognition.draw_face_boxes_from_results(display_frame, display_faces_with_labels)
            else:
                # No faces detected by DeepFace - show positioning guidance
                self._display_distance_feedback(display_frame, "Please position your face in the camera view", False)
        
        except Exception as e:
            print(f"[DEEPFACE ERROR] Detection processing failed: {e}")
        
        return recognized_faces
    
    def _draw_ultra_light_faces(self, frame, faces_with_labels):
        """Draw face detection boxes for ultra light detector"""
        result_frame = frame.copy()
        
        for face in faces_with_labels:
            x, y, w, h = face['position']
            x2, y2 = x + w, y + h
            confidence = face['confidence']
            name = face['name']
            is_best = face.get('is_best', False)
            
            # Color coding
            color = (0, 255, 0) if is_best else (0, 255, 255)  # Green for best, yellow for others
            thickness = 3 if is_best else 2
            
            # Draw bounding box
            cv2.rectangle(result_frame, (x, y), (x2, y2), color, thickness)
            
            # Draw label with background
            label = name
            label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 1)[0]
            
            # Background for text
            cv2.rectangle(result_frame, 
                         (x, y - label_size[1] - 10), 
                         (x + label_size[0], y), 
                         color, -1)
            
            # Text
            cv2.putText(result_frame, label, (x, y - 5), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
        
        # Add performance overlay
        if hasattr(self, 'ultra_light_detector') and self.ultra_light_detector:
            fps = self.ultra_light_detector.get_performance_stats()['fps']
            cv2.putText(result_frame, f"Ultra Light FPS: {fps:.1f}", (10, 30), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        return result_frame

def main():
    """Run the simple kiosk application"""
    app = SimpleKioskApp()
    try:
        app.root.mainloop()
    except KeyboardInterrupt:
        app.quit_app()

if __name__ == "__main__":
    main()
